from __future__ import annotations

import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.common import decimal_value, integer, optional_money as currency
from automations.excel_reader import load_workbook_compatible as load_workbook

TOLERANCE = Decimal("0.01")


# Modelos das ocorrências do Simphony, Fiscal, SEFAZ e do resultado final.
@dataclass(frozen=True)
class SourceEntry:
    value: Decimal
    emission_date: str
    status: str = "Aprovado"


@dataclass(frozen=True)
class CouponRow:
    key: str
    document_type: str
    simphony_date: str
    fiscal_date: str
    sefaz_date: str
    simphony: Decimal | None
    fiscal: Decimal | None
    sefaz: Decimal | None
    simphony_status: str
    status: str

    @property
    def difference(self) -> Decimal:
        if "cancelad" in self.status.casefold() or "cancelamento" in self.status.casefold():
            downstream = [
                abs(value)
                for value in (self.fiscal, self.sefaz)
                if value is not None
            ]
            return max(downstream, default=Decimal())
        values = [
            value
            for value in (self.simphony, self.fiscal, self.sefaz)
            if value is not None
        ]
        return max(values) - min(values) if values else Decimal()

    @property
    def comparable(self) -> bool:
        if "cancelad" in self.status.casefold():
            return True
        return None not in (self.simphony, self.fiscal, self.sefaz)

    @property
    def reference_date(self) -> str:
        return next(
            (
                value
                for value in (self.simphony_date, self.fiscal_date, self.sefaz_date)
                if value != "—"
            ),
            "—",
        )


@dataclass(frozen=True)
class CouponResult:
    rows: list[CouponRow]
    simphony_total: Decimal
    fiscal_total: Decimal
    sefaz_total: Decimal
    cancelled: int
    hotel: str = "Cumbuco"

    def count(self, status: str) -> int:
        return sum(row.status.startswith(status) for row in self.rows)


def date_value(value: Any) -> str:
    if value in (None, "", "NULL"):
        return "—"
    if isinstance(value, (date, datetime)):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    for pattern in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return text


def document_type(key: str) -> str:
    model = key[20:22] if len(key) >= 22 else ""
    return (
        "Cupom (NFC-e)"
        if model == "65"
        else "Nota (NF-e)"
        if model == "55"
        else "Nota (Unknown)"
    )


# Detecta o cabeçalho e lê cada uma das três fontes de cupons e notas.
def workbook_rows(path: Path) -> list[tuple[Any, ...]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def find_header(
    rows: list[tuple[Any, ...]], required: set[str]
) -> tuple[int, tuple[Any, ...]]:
    for index, row in enumerate(rows[:40]):
        if required.issubset(set(row)):
            return index, row
    raise ValueError(f"Cabeçalhos não encontrados: {', '.join(sorted(required))}.")


def read_file(path: Path) -> tuple[str, dict[str, SourceEntry], int]:
    rows = workbook_rows(path)
    headers = {value for row in rows[:40] for value in row if value}
    values: defaultdict[str, Decimal] = defaultdict(Decimal)
    dates: dict[str, str] = {}
    statuses: dict[str, str] = {}
    cancelled = 0

    if {"Chave da NF", "Valor Total NF", "Status"}.issubset(headers):
        source = "Simphony"
        header_i, header = find_header(
            rows, {"Chave da NF", "Valor Total NF", "Status"}
        )
        key_i, value_i, status_i = (
            header.index("Chave da NF"),
            header.index("Valor Total NF"),
            header.index("Status"),
        )
        date_i = header.index("Data")
        for row in rows[header_i + 1 :]:
            if len(row) <= max(key_i, value_i, status_i) or not row[key_i]:
                continue
            status = " ".join(str(row[status_i]).strip().casefold().split())
            key = str(row[key_i]).strip()
            if status not in {"aprovado", "aprovado (c)"}:
                statuses.setdefault(
                    key,
                    "Cancelado" if "cancelad" in status else str(row[status_i]).strip(),
                )
                if "cancelad" in status:
                    cancelled += 1
                values.setdefault(key, decimal_value(row[value_i]))
                dates.setdefault(
                    key, date_value(row[date_i] if len(row) > date_i else None)
                )
                continue
            if statuses.get(key) != "Aprovado":
                values[key] = Decimal()
            statuses[key] = "Aprovado"
            values[key] += decimal_value(row[value_i])
            dates.setdefault(
                key, date_value(row[date_i] if len(row) > date_i else None)
            )
    elif {"Chave", "ValorContabil"}.issubset(headers):
        source = "Fiscal"
        header_i, header = find_header(rows, {"Chave", "ValorContabil"})
        key_i, value_i = header.index("Chave"), header.index("ValorContabil")
        date_i = header.index("DataDocumento")
        cancelled_i = header.index("Cancelado") if "Cancelado" in header else None
        for row in rows[header_i + 1 :]:
            if len(row) <= max(key_i, value_i) or not row[key_i]:
                continue
            if (
                cancelled_i is not None
                and len(row) > cancelled_i
                and str(row[cancelled_i]).lower() == "true"
            ):
                continue
            key = str(row[key_i]).strip()
            values[key] += decimal_value(row[value_i])
            dates.setdefault(
                key, date_value(row[date_i] if len(row) > date_i else None)
            )
    elif {"Chave de acesso", "Valor R$"}.issubset(headers):
        source = "SEFAZ"
        header_i, header = find_header(rows, {"Chave de acesso", "Valor R$"})
        key_i, value_i = header.index("Chave de acesso"), header.index("Valor R$")
        date_i = header.index("Data de emissão")
        for row in rows[header_i + 1 :]:
            if len(row) > max(key_i, value_i) and row[key_i]:
                key = str(row[key_i]).strip()
                values[key] += decimal_value(row[value_i])
                dates.setdefault(
                    key, date_value(row[date_i] if len(row) > date_i else None)
                )
    else:
        raise ValueError(
            f"{path.name}: formato não reconhecido como Simphony, Fiscal ou SEFAZ."
        )
    return (
        source,
        {
            key: SourceEntry(value, dates.get(key, "—"), statuses.get(key, "Aprovado"))
            for key, value in values.items()
        },
        cancelled,
    )


# Compara chave, valor e cancelamento entre Simphony, Fiscal e SEFAZ.
def reconcile(paths: list[Path], hotel: str = "Cumbuco") -> CouponResult:
    if len(paths) != 3:
        raise ValueError("Selecione exatamente as planilhas Simphony, Fiscal e SEFAZ.")
    sources: dict[str, dict[str, SourceEntry]] = {}
    cancelled = 0
    for path in paths:
        source, values, file_cancelled = read_file(path)
        if source in sources:
            raise ValueError(f"Foram selecionados dois arquivos do tipo {source}.")
        sources[source] = values
        cancelled += file_cancelled
    if set(sources) != {"Simphony", "Fiscal", "SEFAZ"}:
        raise ValueError(
            "É necessário selecionar um arquivo Simphony, um Fiscal e um SEFAZ."
        )

    simphony, fiscal, sefaz = sources["Simphony"], sources["Fiscal"], sources["SEFAZ"]
    result_rows: list[CouponRow] = []
    for key in set(simphony) | set(fiscal) | set(sefaz):
        entries = (simphony.get(key), fiscal.get(key), sefaz.get(key))
        values = tuple(entry.value if entry else None for entry in entries)
        missing = [
            name
            for name, value in zip(("Simphony", "Fiscal", "SEFAZ"), entries)
            if value is None
        ]
        if entries[0] is not None and entries[0].status == "Cancelado":
            downstream = (values[1], values[2])
            if all(
                value is None or abs(value) <= TOLERANCE
                for value in downstream
            ):
                status = "Conciliado: cancelado"
            else:
                status = "Divergente: cancelamento"
        elif entries[0] is not None and entries[0].status != "Aprovado":
            status = f"{entries[0].status} no Simphony"
        elif missing:
            status = "Ausente: " + "/".join(missing)
        elif max(values) - min(values) <= TOLERANCE:  # type: ignore[arg-type]
            status = "Conciliado"
        else:
            status = "Divergente: valor"
        source_dates = tuple(entry.emission_date if entry else "—" for entry in entries)
        result_rows.append(
            CouponRow(
                key,
                document_type(key),
                *source_dates,
                *values,
                entries[0].status if entries[0] else "Ausente",
                status,
            )
        )
    result_rows.sort(
        key=lambda row: (row.status == "Conciliado", -row.difference, row.key)
    )
    return CouponResult(
        result_rows,
        sum(
            (
                entry.value
                for entry in simphony.values()
                if entry.status == "Aprovado"
            ),
            Decimal(),
        ),
        sum((entry.value for entry in fiscal.values()), Decimal()),
        sum((entry.value for entry in sefaz.values()), Decimal()),
        cancelled,
        hotel,
    )


# Exporta a conferência de cupons em Excel e PDF.
def save_excel(result: CouponResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    missing = sum(row.status.startswith("Ausente") for row in result.rows)
    summary.append(["Indicador", "Valor"])
    for label, value in (
        ("Hotel", result.hotel),
        ("Total Simphony", float(result.simphony_total)),
        ("Total Fiscal", float(result.fiscal_total)),
        ("Total SEFAZ", float(result.sefaz_total)),
        ("Chaves analisadas", len(result.rows)),
        ("Conciliadas", result.count("Conciliado")),
        ("Divergentes", result.count("Divergente")),
        ("Com integração ausente", missing),
        ("Cupons cancelados ignorados", result.cancelled),
    ):
        summary.append([label, value])
    for cell in summary[1]:
        cell.style = "Headline 4"
    for cell in summary["B"][2:5]:
        cell.number_format = "R$ #,##0.00"
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 20

    headers = [
        "Tipo",
        "Chave fiscal",
        "Data",
        "Simphony - Valor Total NF",
        "Fiscal - Valor Contábil",
        "SEFAZ - Valor",
        "Diferença",
        "Status Simphony",
        "Status",
    ]

    def create_detail_sheet(title: str, rows: list[CouponRow], table_name: str) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row.document_type,
                    row.key,
                    row.reference_date,
                    float(row.simphony) if row.simphony is not None else None,
                    float(row.fiscal) if row.fiscal is not None else None,
                    float(row.sefaz) if row.sefaz is not None else None,
                    float(row.difference) if row.comparable else None,
                    row.simphony_status,
                    row.status,
                ]
            )
        for cell in sheet[1]:
            cell.style = "Headline 4"
        for column in ("D", "E", "F", "G"):
            for cell in sheet[column][1:]:
                cell.number_format = "R$ #,##0.00"
        for column, width in {
            "A": 18,
            "B": 48,
            "C": 14,
            "D": 25,
            "E": 25,
            "F": 22,
            "G": 18,
            "H": 20,
            "I": 24,
        }.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"
        if rows:
            table = ExcelTable(displayName=table_name, ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        else:
            sheet.auto_filter.ref = sheet.dimensions

    create_detail_sheet("Conciliação", result.rows, "TabelaConferenciaCupons")
    create_detail_sheet(
        "Cupons_NFCe",
        [row for row in result.rows if row.document_type == "Cupom (NFC-e)"],
        "TabelaCuponsNFCe",
    )
    create_detail_sheet(
        "Notas_NFe",
        [row for row in result.rows if row.document_type != "Cupom (NFC-e)"],
        "TabelaNotasNFe",
    )
    workbook.save(path)


def save_pdf(result: CouponResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Resumo da conferência dos cupons — {result.hotel}",
    )
    missing = sum(row.status.startswith("Ausente") for row in result.rows)
    data = [
        ["Simphony", "Fiscal", "SEFAZ", "Conciliadas", "Divergentes", "Ausentes"],
        [
            currency(result.simphony_total),
            currency(result.fiscal_total),
            currency(result.sefaz_total),
            integer(result.count("Conciliado")),
            integer(result.count("Divergente")),
            integer(missing),
        ],
    ]
    table = Table(data, colWidths=[39 * mm] * 6)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    document.build(
        [
            Paragraph(
                f"Conferência dos cupons — {result.hotel} — Simphony x Fiscal x SEFAZ",
                styles["Title"],
            ),
            Spacer(1, 6 * mm),
            table,
        ]
    )
