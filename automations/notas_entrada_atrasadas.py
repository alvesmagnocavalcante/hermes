from __future__ import annotations

import re
import unicodedata
import warnings
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

STATE_CODES = {
    "ceara": "CE",
    "sao-paulo": "SP",
    "parana": "PR",
    "distrito-federal": "DF",
    "santa-catarina": "SC",
    "espirito-santo": "ES",
    "alagoas": "AL",
    "pernambuco": "PE",
    "paraiba": "PB",
    "bahia": "BA",
    "rio-grande-do-sul": "RS",
    "minas-gerais": "MG",
    "rio-de-janeiro": "RJ",
    "piaui": "PI",
    "amazonas": "AM",
    "pa": "PA",
    "rio-grande-do-norte": "RN",
}


@dataclass(frozen=True)
class NoteResult:
    key: str
    company: str
    supplier: str
    state: str
    emission_date: date | None
    entry_date: date | None
    days: int | None
    limit: int | None
    launch_status: str
    status: str


@dataclass(frozen=True)
class AnalysisResult:
    rows: list[NoteResult]

    def count(self, status: str) -> int:
        return sum(row.status == status for row in self.rows)


def state_code(value: Any) -> str:
    text = str(value or "").lower()
    match = re.search(r"bandeira-([a-z-]+)\.png", text)
    slug = match.group(1) if match else text.strip().lower()
    return STATE_CODES.get(slug, slug.upper() if len(slug) == 2 else "N/I")


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            continue
    return None


def as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    parsed = as_date(value)
    return datetime.combine(parsed, time.min) if parsed else None


def rounded_days(start: datetime, end: datetime) -> int:
    elapsed = Decimal(str((end - start).total_seconds())) / Decimal(86400)
    return max(0, int(elapsed.quantize(Decimal("1"), rounding=ROUND_HALF_UP)))


def normalized_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^a-z0-9]",
        "",
        "".join(
            character
            for character in text.lower()
            if not unicodedata.combining(character)
        ),
    )


def header_index(headers: tuple[Any, ...], *names: str) -> int:
    available = {
        normalized_header(header): index for index, header in enumerate(headers)
    }
    for name in names:
        index = available.get(normalized_header(name))
        if index is not None:
            return index
    raise ValueError(f"Coluna não encontrada: {' ou '.join(names)}.")


def row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def note_status(days: int, state: str) -> tuple[int, str]:
    alert, limit = (6, 11) if state == "CE" else (20, 30)
    if (state == "CE" and days >= limit) or (state != "CE" and days > limit):
        return limit, "Em atraso"
    if days >= alert:
        return limit, "Alerta"
    return limit, "Em dia"


def open_rows(path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    return workbook, headers, rows


def analyze(paths: list[Path], reference_date: date | None = None) -> AnalysisResult:
    if len(paths) != 2:
        raise ValueError(
            "Selecione exatamente as planilhas Manifesto e Detalhe das notas recebidas."
        )

    manifesto_path: Path | None = None
    detalhe_path: Path | None = None
    for path in paths:
        workbook, headers, _ = open_rows(path)
        workbook.close()
        normalized = {normalized_header(header) for header in headers}
        if normalized_header("Chave Manifesto") in normalized:
            manifesto_path = path
        elif {
            normalized_header("Chave"),
            normalized_header("Data de Entrada"),
        } <= normalized:
            detalhe_path = path
    if not manifesto_path or not detalhe_path:
        raise ValueError(
            "Não foi possível identificar uma planilha Manifesto e uma planilha Detalhe."
        )

    workbook, headers, rows = open_rows(manifesto_path)
    key_i = header_index(headers, "Chave Manifesto")
    emission_i = header_index(headers, "Data Emissão", "Data da Emissão")
    state_i = header_index(headers, "Estado")
    company_i = header_index(headers, "Empresa")
    supplier_i = header_index(headers, "Fornecedor", "Razão Social")
    required_max_i = max(key_i, emission_i, company_i)
    manifest: dict[tuple[str, str], tuple[datetime | None, str, str, str]] = {}
    try:
        for row in rows:
            if len(row) <= required_max_i or not row[key_i]:
                continue
            company = str(row[company_i] or "").strip()
            key = str(row[key_i]).strip()
            manifest[(normalized_header(company), key)] = (
                as_datetime(row[emission_i]),
                state_code(row_value(row, state_i)),
                company,
                str(row_value(row, supplier_i) or ""),
            )
    finally:
        workbook.close()

    workbook, headers, rows = open_rows(detalhe_path)
    key_i = header_index(headers, "Chave")
    company_i = header_index(headers, "Empresa")
    entry_i = header_index(headers, "Data de Entrada")
    max_i = max(key_i, company_i, entry_i)
    entries: dict[tuple[str, str], datetime | None] = {}
    try:
        for row in rows:
            if len(row) <= max_i or not row[key_i]:
                continue
            key = (
                normalized_header(row[company_i]),
                str(row[key_i]).strip(),
            )
            entry = as_datetime(row[entry_i])
            previous = entries.get(key)
            if previous is None or (entry is not None and entry < previous):
                entries[key] = entry
    finally:
        workbook.close()

    today = datetime.combine(reference_date or date.today(), time.min)
    results: list[NoteResult] = []
    for company_key, (emission, state, company, supplier) in manifest.items():
        key = company_key[1]
        entry = entries.get(company_key)
        launch_status = "Lançada" if entry else "Não lançada"
        if not emission:
            results.append(
                NoteResult(
                    key,
                    company,
                    supplier,
                    state,
                    None,
                    entry.date() if entry else None,
                    None,
                    None,
                    f"{launch_status} • emissão ausente",
                    "Alerta",
                )
            )
            continue
        reference = entry or today
        days = rounded_days(emission, reference)
        limit, status = note_status(days, state)
        results.append(
            NoteResult(
                key,
                company,
                supplier,
                state,
                emission.date(),
                entry.date() if entry else None,
                days,
                limit,
                launch_status,
                status,
            )
        )
    order = {"Em atraso": 0, "Alerta": 1, "Em dia": 2}
    results.sort(key=lambda item: (order[item.status], -(item.days or 0), item.key))
    return AnalysisResult(results)


def date_text(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else "—"


def save_excel(result: AnalysisResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Quantidade"])
    summary.append(["Notas analisadas", len(result.rows)])
    summary.append(["Em dia", result.count("Em dia")])
    summary.append(["Em alerta", result.count("Alerta")])
    summary.append(["Em atraso", result.count("Em atraso")])
    summary.append(["Não lançadas", sum(row.entry_date is None for row in result.rows)])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 16
    for cell in summary[1]:
        cell.style = "Headline 4"

    details = workbook.create_sheet("Análise")
    details.append(
        [
            "Chave",
            "Empresa",
            "Fornecedor",
            "UF",
            "Emissão",
            "Entrada",
            "Dias",
            "Limite para atraso",
            "Lançamento",
            "Situação",
        ]
    )
    for row in result.rows:
        details.append(
            [
                row.key,
                row.company,
                row.supplier,
                row.state,
                row.emission_date,
                row.entry_date,
                row.days,
                row.limit,
                row.launch_status,
                row.status,
            ]
        )
    for cell in details[1]:
        cell.style = "Headline 4"
    for column in ("E", "F"):
        for cell in details[column][1:]:
            cell.number_format = "DD/MM/YYYY"
    for column, width in {
        "A": 48,
        "B": 22,
        "C": 45,
        "D": 8,
        "E": 14,
        "F": 14,
        "G": 10,
        "H": 18,
        "I": 24,
        "J": 16,
    }.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    workbook.save(path)


def save_pdf(result: AnalysisResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Resumo de notas fiscais em atraso",
    )
    data = [
        ["Analisadas", "Em dia", "Em alerta", "Em atraso", "Não lançadas"],
        [
            len(result.rows),
            result.count("Em dia"),
            result.count("Alerta"),
            result.count("Em atraso"),
            sum(row.entry_date is None for row in result.rows),
        ],
    ]
    table = Table(data, colWidths=[45 * mm] * 5)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    document.build(
        [
            Paragraph(
                "Notas fiscais de entrada de mercadoria em atraso", styles["Title"]
            ),
            Spacer(1, 6 * mm),
            table,
        ]
    )
