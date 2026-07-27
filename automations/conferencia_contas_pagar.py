from __future__ import annotations

import re
import unicodedata
import warnings
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

TOLERANCE = Decimal("0.01")

ENTITY_GROUPS = (
    ("CVC", ("CVC",)),
    ("DECOLAR / DESPEGAR", ("DECOLAR", "DESPEGAR")),
    ("BESTBUY", ("BESTBUY",)),
    ("BRT", ("BRT",)),
    ("DIVERSA", ("DIVERSA",)),
    ("INTEREP", ("INTEREP",)),
    ("ORINTER", ("ORINTER",)),
    ("PRIMETOUR", ("PRIMETOUR",)),
    ("TREND", ("TREND",)),
    ("VIAGENS PROMO", ("VIAGENSPROMO",)),
    ("VISUAL", ("VISUAL",)),
    ("STONE LINK", ("STONELINK", "LINKSTONE")),
    ("REDE HIPERCARD", ("REDEHIPERCARD", "REDECARDHIPERCARD")),
)


@dataclass(frozen=True)
class EntityRow:
    category: str
    name: str
    accounting: Decimal
    financial: Decimal

    @property
    def difference(self) -> Decimal:
        return self.financial - self.accounting

    @property
    def status(self) -> str:
        return "Conciliado" if abs(self.difference) <= TOLERANCE else "Divergente"


@dataclass(frozen=True)
class Check:
    name: str
    financial: Decimal
    accounting: Decimal

    @property
    def difference(self) -> Decimal:
        return self.financial - self.accounting

    @property
    def status(self) -> str:
        return "Conciliado" if abs(self.difference) <= TOLERANCE else "Divergente"


@dataclass(frozen=True)
class PayablesResult:
    entities: list[EntityRow]
    suppliers: Check
    advances: Check
    taxes: tuple[Check, Check, Check]
    hotel: str = "Cumbuco"


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char)).upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def entity_group(value: Any) -> tuple[str, str]:
    normalized = normalize(value)
    for label, aliases in ENTITY_GROUPS:
        if any(alias in normalized for alias in aliases):
            return normalize(label), label
    return normalized, str(value).strip()


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal()
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(str(value).replace(".", "").replace(",", "."))


def currency(value: Decimal) -> str:
    text = f"R$ {value:,.2f}"
    return text.replace(",", "_").replace(".", ",").replace("_", ".")


def integer(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def read(path: Path) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        rows = list(sheet.iter_rows(values_only=True))
        return rows[0], rows[1:]
    finally:
        workbook.close()


def identify(
    paths: list[Path],
) -> dict[str, tuple[tuple[Any, ...], list[tuple[Any, ...]]]]:
    files: dict[str, tuple[tuple[Any, ...], list[tuple[Any, ...]]]] = {}
    for path in paths:
        header, rows = read(path)
        headers = set(header)
        name = normalize(path.name)
        if {"DescricaoSubconta", "Saldo"}.issubset(headers):
            key = "balancete_adto" if "ADTO" in name else "balancete_fornecedor"
        elif {"Fornecedor", "Saldo", "DescContaContabil"}.issubset(headers):
            key = "posicao_fornecedor"
        elif {"NomeFornecedor", "Saldo", "ValorTotalAdiantado"}.issubset(headers):
            key = "adiantamentos"
        elif {"IdAgregado", "Valor", "TratamentoFiscal"}.issubset(headers):
            tax = next((tax for tax in ("IRRF", "CSRF", "ISS") if tax in name), None)
            if tax is None:
                raise ValueError(
                    f"{path.name}: informe IRRF, CSRF ou ISS no nome do arquivo."
                )
            key = f"agregado_{tax.lower()}"
        elif {"DescricaoConta", "Movimento", "Historico"}.issubset(headers):
            key = "razao_impostos"
        else:
            raise ValueError(
                f"{path.name}: arquivo não reconhecido para esta conferência."
            )
        if key in files:
            raise ValueError(f"Dois arquivos foram identificados como {key}.")
        files[key] = (header, rows)
    expected = {
        "balancete_adto",
        "balancete_fornecedor",
        "posicao_fornecedor",
        "adiantamentos",
        "agregado_irrf",
        "agregado_csrf",
        "agregado_iss",
        "razao_impostos",
    }
    if set(files) != expected:
        raise ValueError("Selecione os oito arquivos da Atividade 9.")
    return files


def grouped(
    data, name_column: str, value_column: str
) -> dict[str, tuple[str, Decimal]]:
    header, rows = data
    name_i, value_i = header.index(name_column), header.index(value_column)
    sums: defaultdict[str, Decimal] = defaultdict(Decimal)
    labels: dict[str, str] = {}
    for row in rows:
        if len(row) <= max(name_i, value_i) or row[name_i] in (None, "", "NULL"):
            continue
        key, label = entity_group(row[name_i])
        labels.setdefault(key, label)
        sums[key] += decimal_value(row[value_i])
    return {key: (labels[key], abs(value)) for key, value in sums.items()}


def entities(category: str, accounting, financial) -> list[EntityRow]:
    result: list[EntityRow] = []
    accounting_left = dict(accounting)
    financial_left = dict(financial)

    for key in set(accounting_left) & set(financial_left):
        accounting_name, accounting_value = accounting_left.pop(key)
        _, financial_value = financial_left.pop(key)
        result.append(
            EntityRow(category, accounting_name, accounting_value, financial_value)
        )

    candidates = sorted(
        [
            (
                SequenceMatcher(None, accounting_key, financial_key).ratio(),
                accounting_key,
                financial_key,
            )
            for accounting_key, (_, accounting_value) in accounting_left.items()
            for financial_key, (_, financial_value) in financial_left.items()
            if accounting_value != 0
            and abs(accounting_value - financial_value) <= TOLERANCE
        ],
        reverse=True,
    )
    for _, accounting_key, financial_key in candidates:
        if accounting_key not in accounting_left or financial_key not in financial_left:
            continue
        accounting_name, accounting_value = accounting_left.pop(accounting_key)
        financial_name, financial_value = financial_left.pop(financial_key)
        label = (
            accounting_name
            if accounting_name == financial_name
            else f"{accounting_name} / {financial_name}"
        )
        result.append(EntityRow(category, label, accounting_value, financial_value))

    result.extend(
        EntityRow(category, name, value, Decimal())
        for name, value in accounting_left.values()
    )
    result.extend(
        EntityRow(category, name, Decimal(), value)
        for name, value in financial_left.values()
    )
    return result


def absolute_total(data, column: str) -> Decimal:
    header, rows = data
    index = header.index(column)
    return sum(
        (abs(decimal_value(row[index])) for row in rows if len(row) > index), Decimal()
    )


def analyze(paths: list[Path], hotel: str = "Cumbuco") -> PayablesResult:
    if len(paths) != 8:
        raise ValueError("Selecione exatamente os oito arquivos da Atividade 9.")
    files = identify(paths)
    supplier_accounting = grouped(
        files["balancete_fornecedor"], "DescricaoSubconta", "Saldo"
    )
    supplier_financial = grouped(files["posicao_fornecedor"], "Fornecedor", "Saldo")
    advance_accounting = grouped(files["balancete_adto"], "DescricaoSubconta", "Saldo")
    advance_financial = grouped(files["adiantamentos"], "NomeFornecedor", "Saldo")
    rows = entities("Fornecedores", supplier_accounting, supplier_financial)
    rows.extend(entities("Adiantamentos", advance_accounting, advance_financial))
    rows.sort(
        key=lambda row: (row.status == "Conciliado", -abs(row.difference), row.name)
    )

    tax_financial = {
        "IRRF": absolute_total(files["agregado_irrf"], "Valor"),
        "CSRF": absolute_total(files["agregado_csrf"], "Valor"),
        "ISS": absolute_total(files["agregado_iss"], "Valor"),
    }
    header, ledger_rows = files["razao_impostos"]
    name_i, movement_i = header.index("DescricaoConta"), header.index("Movimento")
    tax_accounting = {"IRRF": Decimal(), "CSRF": Decimal(), "ISS": Decimal()}
    for row in ledger_rows:
        if len(row) <= max(name_i, movement_i):
            continue
        movement = decimal_value(row[movement_i])
        if movement >= 0:
            continue
        account = normalize(row[name_i])
        tax = (
            "IRRF"
            if "IRRF" in account
            else "CSRF"
            if "PISCOFINSCSLL" in account
            else "ISS"
            if "ISSRETIDO" in account
            else None
        )
        if tax:
            tax_accounting[tax] += abs(movement)

    supplier_accounting_total = sum(
        (value for _, value in supplier_accounting.values()), Decimal()
    )
    supplier_financial_total = sum(
        (value for _, value in supplier_financial.values()), Decimal()
    )
    advance_accounting_total = sum(
        (value for _, value in advance_accounting.values()), Decimal()
    )
    advance_financial_total = sum(
        (value for _, value in advance_financial.values()), Decimal()
    )
    return PayablesResult(
        rows,
        Check("Fornecedores", supplier_financial_total, supplier_accounting_total),
        Check("Adiantamentos", advance_financial_total, advance_accounting_total),
        tuple(
            Check(tax, tax_financial[tax], tax_accounting[tax])
            for tax in ("IRRF", "CSRF", "ISS")
        ),
        hotel,
    )


def save_excel(result: PayablesResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Conferência do Contas a Pagar"])
    summary.append(["Hotel", result.hotel])
    summary.append([])
    summary.append(
        ["Conferência", "Financeiro", "Contabilidade", "Diferença", "Status"]
    )
    for check in (result.suppliers, result.advances, *result.taxes):
        summary.append(
            [
                check.name,
                float(check.financial),
                float(check.accounting),
                float(check.difference),
                check.status,
            ]
        )
    summary["A1"].style = "Title"
    summary["A2"].style = "Headline 4"
    for cell in summary[4]:
        cell.style = "Headline 4"
    for column in ("B", "C", "D"):
        for cell in summary[column][4:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {"A": 25, "B": 20, "C": 20, "D": 18, "E": 16}.items():
        summary.column_dimensions[column].width = width
    details = workbook.create_sheet("Fornecedores e Adiantamentos")
    details.append(["Conferência do Contas a Pagar"])
    details.append(["Hotel", result.hotel])
    details.append([])
    details.append(
        [
            "Tipo",
            "Fornecedor / Subconta",
            "Contabilidade",
            "Financeiro",
            "Diferença",
            "Status",
        ]
    )
    for row in result.entities:
        details.append(
            [
                row.category,
                row.name,
                float(row.accounting),
                float(row.financial),
                float(row.difference),
                row.status,
            ]
        )
    details["A1"].style = "Title"
    details["A2"].style = "Headline 4"
    for cell in details[4]:
        cell.style = "Headline 4"
    for column in ("C", "D", "E"):
        for cell in details[column][4:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {"A": 18, "B": 50, "C": 20, "D": 20, "E": 18, "F": 16}.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A5"
    details.auto_filter.ref = f"A4:F{details.max_row}"
    workbook.save(path)


def save_pdf(result: PayablesResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Conferência do Contas a Pagar",
    )
    data = [["Conferência", "Financeiro", "Contabilidade", "Diferença", "Status"]]
    for check in (result.suppliers, result.advances, *result.taxes):
        data.append(
            [
                check.name,
                currency(check.financial),
                currency(check.accounting),
                currency(check.difference),
                check.status,
            ]
        )
    table = Table(data, colWidths=[48 * mm, 44 * mm, 44 * mm, 42 * mm, 35 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    document.build(
        [
            Paragraph("Conferência do Contas a Pagar", styles["Title"]),
            Spacer(1, 6 * mm),
            table,
        ]
    )
