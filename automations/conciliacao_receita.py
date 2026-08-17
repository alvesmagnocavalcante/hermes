from __future__ import annotations

import warnings
from calendar import monthrange
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from itertools import combinations
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.common import decimal_value, integer, money as currency
from automations.excel_reader import load_workbook_compatible as load_workbook

TOLERANCE = Decimal("0.01")


# Estruturas que representam cada dia conciliado e os totais do período.
@dataclass(frozen=True)
class ReconciliationRow:
    business_date: date
    cmflex: Decimal
    opera: Decimal
    identification: str = "—"

    @property
    def difference(self) -> Decimal:
        return self.cmflex - self.opera

    @property
    def status(self) -> str:
        return "Conciliado" if abs(self.difference) <= TOLERANCE else "Divergente"


@dataclass(frozen=True)
class ReconciliationResult:
    rows: list[ReconciliationRow]
    cmflex_total: Decimal
    opera_total: Decimal

    @property
    def difference(self) -> Decimal:
        return self.cmflex_total - self.opera_total

    @property
    def reconciled(self) -> int:
        return sum(row.status == "Conciliado" for row in self.rows)


# Converte datas e valores e extrai os movimentos das planilhas de origem.
def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(
        "".join(char for char in text if not unicodedata.combining(char))
        .upper()
        .split()
    )


def date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value if value.year >= 1900 else None
    text = str(value or "").strip()
    for pattern in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d-%b-%y",
        "%Y-%m-%d",
    ):
        try:
            parsed = datetime.strptime(text, pattern).date()
            return parsed if parsed.year >= 1900 else None
        except ValueError:
            continue
    return None


def ignored_opera_transaction(description: Any, hotel: str) -> bool:
    transaction = normalize(description)
    return normalize(hotel) == "MAGNA" and transaction in {
        "AJUSTE TAXA DE TURISMO",
        "TAXA DE TURISMO",
    }


def read_values(
    path: Path, hotel: str
) -> tuple[str, dict[date, Decimal], dict[date, dict[str, Decimal]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        # Alguns relatórios exportados pelos sistemas não gravam a dimensão da
        # planilha. No modo somente leitura, calculate_dimension() lança
        # "Worksheet is unsized" antes que seja possível percorrer as linhas.
        sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(next(rows))

        if "Movimento" in headers and "Documento" in headers:
            source, date_name, value_name, multiplier = (
                "Contabilidade",
                "DataLancamento",
                "Movimento",
                Decimal(-1),
            )
        elif "CASHIER_DEBIT" in headers and "TRX_NO" in headers:
            source, date_name, value_name, multiplier = (
                "Opera",
                "BUSINESS_FORMAT_DATE",
                "CASHIER_DEBIT",
                Decimal(1),
            )
        else:
            raise ValueError(
                f"{path.name}: não foram encontradas as colunas esperadas "
                "(Movimento/DataLancamento ou "
                "CASHIER_DEBIT/BUSINESS_FORMAT_DATE)."
            )

        if date_name not in headers:
            raise ValueError(f"{path.name}: a coluna {date_name} não foi encontrada.")
        date_index, value_index = headers.index(date_name), headers.index(value_name)
        description_index = (
            headers.index("TRX_DESC")
            if source == "Opera" and "TRX_DESC" in headers
            else None
        )
        values: defaultdict[date, Decimal] = defaultdict(Decimal)
        transaction_totals: defaultdict[date, defaultdict[str, Decimal]] = defaultdict(
            lambda: defaultdict(Decimal)
        )
        for row in rows:
            if date_index >= len(row) or value_index >= len(row):
                continue
            business_date = date_value(row[date_index])
            if business_date is None:
                continue
            description = (
                row[description_index]
                if description_index is not None and description_index < len(row)
                else None
            )
            if description_index is not None and ignored_opera_transaction(
                description, hotel
            ):
                continue
            amount = decimal_value(row[value_index]) * multiplier
            values[business_date] += amount
            if source == "Opera" and description not in (None, ""):
                transaction_totals[business_date][str(description).strip()] += amount
        return source, dict(values), {
            day: dict(transactions)
            for day, transactions in transaction_totals.items()
        }
    finally:
        workbook.close()


# Consolida Opera e Contabilidade por data e calcula as divergências.
def reconcile(paths: list[Path], hotel: str = "Cumbuco") -> ReconciliationResult:
    if len(paths) != 2:
        raise ValueError(
            "Selecione exatamente o arquivo da Contabilidade e o arquivo do Opera."
        )

    sources: dict[str, dict[date, Decimal]] = {}
    opera_transactions: dict[date, dict[str, Decimal]] = {}
    for path in paths:
        source, values, identified = read_values(path, hotel)
        if source in sources:
            raise ValueError(f"Foram selecionados dois arquivos do tipo {source}.")
        sources[source] = values
        if source == "Opera":
            opera_transactions = identified
    if set(sources) != {"Contabilidade", "Opera"}:
        raise ValueError(
            "É necessário selecionar um arquivo da Contabilidade e um do Opera."
        )

    cmflex, opera = sources["Contabilidade"], sources["Opera"]

    def matching_transactions(
        business_date: date, difference: Decimal
    ) -> list[tuple[str, Decimal]]:
        target = -difference
        if target <= TOLERANCE:
            return []
        candidates = [
            (description, value)
            for description, value in opera_transactions.get(
                business_date, {}
            ).items()
            if value > TOLERANCE and value <= target + TOLERANCE
        ]
        for size in range(1, min(3, len(candidates)) + 1):
            for match in combinations(candidates, size):
                if abs(sum((value for _, value in match), Decimal()) - target) <= TOLERANCE:
                    return list(match)
        return []

    def identification(business_date: date) -> str:
        difference = cmflex.get(business_date, Decimal()) - opera.get(
            business_date, Decimal()
        )
        transactions = matching_transactions(business_date, difference)
        if transactions:
            return "; ".join(
                f"{description} (Journal): {currency(value)}"
                for description, value in transactions
            )
        return (
            "—"
            if abs(difference) <= TOLERANCE
            else "Diferença sem identificação automática"
        )

    periods: defaultdict[tuple[int, int], int] = defaultdict(int)
    for business_date in set(cmflex) | set(opera):
        periods[(business_date.year, business_date.month)] += 1
    if not periods:
        raise ValueError("Os arquivos não contêm datas válidas para conciliação.")
    year, month = max(periods, key=lambda period: periods[period])
    business_dates = [
        date(year, month, day) for day in range(1, monthrange(year, month)[1] + 1)
    ]

    rows = [
        ReconciliationRow(
            business_date,
            cmflex.get(business_date, Decimal()),
            opera.get(business_date, Decimal()),
            identification(business_date),
        )
        for business_date in business_dates
    ]
    return ReconciliationResult(
        rows,
        sum((row.cmflex for row in rows), Decimal()),
        sum((row.opera for row in rows), Decimal()),
    )


# Exporta a conciliação consolidada para Excel e PDF.
def save_excel_result(result: ReconciliationResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Valor"])
    summary.append(["Total Contabilidade (Movimento)", float(result.cmflex_total)])
    summary.append(["Total Opera (CASHIER_DEBIT)", float(result.opera_total)])
    summary.append(["Diferença", float(result.difference)])
    summary.append(["Transações conciliadas", result.reconciled])
    summary.append(["Transações divergentes", len(result.rows) - result.reconciled])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 20
    for cell in summary[1]:
        cell.style = "Headline 4"
    for cell in summary["B"][1:4]:
        cell.number_format = "R$ #,##0.00"

    details = workbook.create_sheet("Conciliação")
    details.append(
        [
            "Data",
            "Saldo Opera",
            "Saldo Contabilidade",
            "Diferença",
            "Identificação",
            "Status",
        ]
    )
    for row in result.rows:
        details.append(
            [
                row.business_date,
                float(row.opera),
                float(row.cmflex),
                float(row.difference),
                row.identification,
                row.status,
            ]
        )
    for cell in details[1]:
        cell.style = "Headline 4"
    for cell in details["A"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for column in ("B", "C", "D"):
        for cell in details[column][1:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {
        "A": 16,
        "B": 20,
        "C": 22,
        "D": 18,
        "E": 48,
        "F": 16,
    }.items():
        details.column_dimensions[column].width = width
    details.auto_filter.ref = details.dimensions
    details.freeze_panes = "A2"
    workbook.save(path)


def save_pdf_result(result: ReconciliationResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Conciliação de Receita",
    )
    divergent = len(result.rows) - result.reconciled
    summary_data = [
        [
            "Total Contabilidade",
            "Total Opera",
            "Diferença",
            "Conciliadas",
            "Divergentes",
        ],
        [
            currency(result.cmflex_total),
            currency(result.opera_total),
            currency(result.difference),
            integer(result.reconciled),
            integer(divergent),
        ],
    ]
    summary = Table(
        summary_data, colWidths=[48 * mm, 48 * mm, 48 * mm, 35 * mm, 35 * mm]
    )
    summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    document.build(
        [
            Paragraph(
                "Conciliação de Receita — Contabilidade x Opera", styles["Title"]
            ),
            Spacer(1, 5 * mm),
            summary,
            Spacer(1, 5 * mm),
        ]
    )
