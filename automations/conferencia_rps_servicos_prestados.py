from __future__ import annotations

import re
import unicodedata
import warnings
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

DEFAULT_SERVICE_CODES = {
    "1000",
    "1003",
    "1004",
    "1005",
    "1022",
    "1122",  # diárias e ajustes
    "3010",
    "3015",
    "3115",  # upselling, toalha e ajuste
    "6000",
    "6003",  # lavanderia e avaria
    "9999",  # diárias de encerramento
}

TAIBA_SERVICE_CODES = {
    "1001",
    "1008",
    "1012",
    "1018",
    "1025",
    "2418",
    "3510",
    "5020",
    "5024",
    "6014",
    "6017",
    "6018",
    "6019",
    "6023",
    "9980",
}

CUMBUCO_SERVICE_CODES = {
    "1000",
    "1002",
    "1004",
    "1008",
    "1010",
    "1011",
    "1022",
    "1023",
    "1111",
    "3010",
    "3012",
    "3013",
    "3014",
    "3017",
    "5002",
    "5004",
    "5005",
    "5007",
    "5008",
    "5009",
    "5011",
    "6000",
    "9999",
}

CHARME_SERVICE_CODES = CUMBUCO_SERVICE_CODES | {
    "1001",
    "1009",
    "1019",
    "3011",
    "5081",
}

SERVICE_CODE_PROFILES = {
    "Padrão": DEFAULT_SERVICE_CODES,
    "Taíba": TAIBA_SERVICE_CODES,
    "Cumbuco": CUMBUCO_SERVICE_CODES,
    "Charme": CHARME_SERVICE_CODES,
}


@dataclass(frozen=True)
class OperaRPS:
    rps: str
    issue_date: date
    room: str
    guest: str
    value: Decimal
    services: str


@dataclass(frozen=True)
class FiscalRPS:
    rps: str
    issue_date: date | None
    nfse: str
    customer: str
    value: Decimal
    status: str


@dataclass(frozen=True)
class CityRPS:
    rps: str
    issue_date: date | None
    nfse: str
    customer: str
    value: Decimal
    status: str


@dataclass(frozen=True)
class RPSResult:
    rps: str
    opera_date: date | None
    room: str
    customer: str
    services: str
    opera_value: Decimal | None
    fiscal_value: Decimal | None
    city_value: Decimal | None
    fiscal_nfse: str
    city_nfse: str
    fiscal_status: str
    city_status: str
    status: str
    detail: str

    @property
    def reconciled(self) -> bool:
        return self.status == "Conciliado"

    @property
    def incomplete(self) -> bool:
        return self.status == "Fora do período do Fiscal"

    @property
    def cancelled(self) -> bool:
        return self.status == "Cancelado"

    @property
    def difference(self) -> Decimal | None:
        values = [
            value
            for value in (self.opera_value, self.fiscal_value, self.city_value)
            if value is not None
        ]
        return max(values) - min(values) if len(values) > 1 else None


@dataclass(frozen=True)
class AnalysisResult:
    rows: list[RPSResult]
    fiscal_start: date
    fiscal_end: date
    files: tuple[str, str, str]
    service_profile: str

    @property
    def reconciled(self) -> int:
        return sum(row.reconciled for row in self.rows)

    @property
    def incomplete(self) -> int:
        return sum(row.incomplete for row in self.rows)

    @property
    def cancelled(self) -> int:
        return sum(row.cancelled for row in self.rows)

    @property
    def issues(self) -> int:
        return len(self.rows) - self.reconciled - self.incomplete - self.cancelled


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^A-Z0-9]",
        "",
        "".join(char for char in text if not unicodedata.combining(char)).upper(),
    )


def decimal_value(value: Any) -> Decimal:
    if value in (None, "", "NULL", "-"):
        return Decimal()
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal()


def money(value: Decimal | None) -> str:
    if value is None:
        return "—"
    return f"R$ {value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def identifier(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d-%b-%y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text.upper(), pattern).date()
        except ValueError:
            continue
    return None


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self.table: list[list[str]] | None = None
        self.row: list[str] | None = None
        self.cell: list[str] | None = None

    def handle_starttag(self, tag: str, _attrs) -> None:
        tag = tag.lower()
        if tag == "table" and self.table is None:
            self.table = []
        elif tag == "tr" and self.table is not None:
            self.row = []
        elif tag in {"td", "th"} and self.row is not None:
            self.cell = []
        elif tag == "br" and self.cell is not None:
            self.cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self.cell is not None:
            self.cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self.cell is not None and self.row is not None:
            self.row.append(re.sub(r"\s+", " ", "".join(self.cell)).strip())
            self.cell = None
        elif tag == "tr" and self.row is not None and self.table is not None:
            if any(self.row):
                self.table.append(self.row)
            self.row = None
        elif tag == "table" and self.table is not None:
            if self.table:
                self.tables.append(self.table)
            self.table = None


def _html_sheet_rows(path: Path) -> tuple[list[Any], list[dict[str, Any]]]:
    content = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            html = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    parser = _HtmlTableParser()
    parser.feed(html)
    signatures = (
        {"NUMERODOCUMENTO", "VALORDOCUMENTO", "SITUACAODOCUMENTO"},
        {"NUMERORPS", "VALORDOSSERVICOS", "STATUSDOC"},
        {"NORPS", "VALORDANFE", "SITUACAO"},
        {"NUMERO", "DATAEMISSAO", "VALORSERVICOS", "NOMETOMADOR"},
        {"NUMRPS", "VALOR", "SITUACAONOTA"},
    )
    for table in parser.tables:
        for index, candidate in enumerate(table):
            keys = {normalize(value) for value in candidate}
            if any(signature.issubset(keys) for signature in signatures):
                headers = candidate
                rows = [
                    dict(zip(headers, row)) for row in table[index + 1 :] if any(row)
                ]
                return headers, rows
    raise ValueError(
        f"As colunas esperadas não foram encontradas no arquivo {path.name}."
    )


def _sheet_rows(path: Path) -> tuple[list[Any], list[dict[str, Any]]]:
    if path.suffix.lower() == ".xls":
        header = path.read_bytes()[:1024].lower()
        if b"<html" in header or b"<!doctype html" in header:
            return _html_sheet_rows(path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        iterator = sheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        return headers, [dict(zip(headers, row)) for row in iterator]
    finally:
        workbook.close()


def identify_file(path: Path) -> str:
    if path.suffix.lower() == ".xml":
        try:
            return (
                "opera"
                if ET.parse(path).getroot().tag == "FOLIO_DETAILS"
                else "unknown"
            )
        except ET.ParseError:
            return "unknown"
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"}:
        return "unknown"
    headers, _ = _sheet_rows(path)
    keys = {normalize(header) for header in headers}
    if {"NUMERODOCUMENTO", "VALORDOCUMENTO", "SITUACAODOCUMENTO"}.issubset(keys):
        return "fiscal"
    if (
        {"NUMERORPS", "VALORDOSSERVICOS", "STATUSDOC"}.issubset(keys)
        or {"NORPS", "VALORDANFE", "SITUACAO"}.issubset(keys)
        or {"NUMERO", "DATAEMISSAO", "VALORSERVICOS", "NOMETOMADOR"}.issubset(keys)
        or {"NUMRPS", "VALOR", "SITUACAONOTA"}.issubset(keys)
    ):
        return "city"
    return "unknown"


def read_opera(
    path: Path, service_codes: set[str] = DEFAULT_SERVICE_CODES
) -> dict[str, OperaRPS]:
    result = {}
    for bill in ET.parse(path).getroot().findall(".//G_BILL_NO"):
        if (bill.findtext("FOLIO_TYPE") or "").upper() != "NOTA":
            continue
        total = Decimal()
        descriptions = set()
        for transaction in bill.findall(".//G_TRX_NO"):
            if transaction.findtext("TRX_CODE") not in service_codes:
                continue
            total += decimal_value(transaction.findtext("FT_DEBIT"))
            description = re.sub(
                r"\s+", " ", transaction.findtext("TRANSACTION_DESCRIPTION") or ""
            ).strip()
            if description:
                descriptions.add(description)
        if not total:
            continue
        rps = str(bill.findtext("BILL_NO") or "").strip()
        issued = parse_date(bill.findtext("BILL_GENERATION_DATE_CHAR"))
        if not rps or not issued:
            continue
        result[rps] = OperaRPS(
            rps,
            issued,
            str(bill.findtext("ROOM") or "").strip().zfill(4),
            str(bill.findtext("DISPLAY_NAME") or "").strip(),
            abs(total).quantize(Decimal("0.01")),
            ", ".join(sorted(descriptions)),
        )
    return result


def read_fiscal(path: Path) -> dict[str, FiscalRPS]:
    _, rows = _sheet_rows(path)
    result = {}
    for row in rows:
        rps = identifier(row.get("NumeroDocumento"))
        if not rps or rps == "NULL":
            continue
        result[rps] = FiscalRPS(
            rps,
            parse_date(row.get("DataEmissao")),
            identifier(row.get("NumeroNFSe")),
            str(row.get("NomeTomadorPrestador") or "").strip(),
            decimal_value(row.get("ValorDocumento"))
            .copy_abs()
            .quantize(Decimal("0.01")),
            str(row.get("SituacaoDocumento") or "").strip(),
        )
    return result


def read_city(
    path: Path, fiscal: dict[str, FiscalRPS] | None = None
) -> dict[str, CityRPS]:
    _, rows = _sheet_rows(path)
    result = {}
    fiscal_by_nfse = {
        item.nfse: item.rps for item in (fiscal or {}).values() if item.nfse
    }
    for row in rows:
        values = {normalize(key): value for key, value in row.items()}
        rps = identifier(
            values.get("NUMERORPS") or values.get("NORPS") or values.get("NUMRPS")
        )
        nfse = identifier(values.get("NUMERO") or values.get("NONOTA"))
        if not rps:
            rps = fiscal_by_nfse.get(nfse, "")
        if not rps and not nfse:
            continue
        key = rps or f"NFS-e {nfse}"
        result[key] = CityRPS(
            rps,
            parse_date(values.get("DATA") or values.get("DATADEEMISSAO")),
            nfse,
            str(
                values.get("RAZAOSOCIALNOMEDOTOMADOR")
                or values.get("TOMADORDESERVICO")
                or values.get("NOMETOMADOR")
                or ""
            ).strip(),
            decimal_value(
                values.get("VALORDOSSERVICOS")
                or values.get("VALORDANFE")
                or values.get("VALORSERVICOS")
                or values.get("VALOR")
            )
            .copy_abs()
            .quantize(Decimal("0.01")),
            str(
                values.get("STATUSDOC")
                or values.get("SITUACAONOTA")
                or values.get("SITUACAO")
                or "Emitida"
            ).strip(),
        )
    return result


def analyze(paths: list[Path]) -> AnalysisResult:
    if len(paths) != 3:
        raise ValueError(
            "Selecione um XML do Opera, uma planilha Fiscal e uma planilha da Prefeitura."
        )
    identified = [(path, identify_file(path)) for path in paths]
    unknown = [path.name for path, kind in identified if kind == "unknown"]
    if unknown:
        raise ValueError(
            f"Arquivo não reconhecido pelo conteúdo: {', '.join(unknown)}."
        )
    grouped = {
        kind: [path for path, current in identified if current == kind]
        for kind in ("opera", "fiscal", "city")
    }
    if any(len(items) != 1 for items in grouped.values()):
        raise ValueError(
            "Envie exatamente um arquivo de cada fonte: Opera, Fiscal e Prefeitura."
        )
    fiscal = read_fiscal(grouped["fiscal"][0])
    city = read_city(grouped["city"][0], fiscal)
    candidates = {
        profile: read_opera(grouped["opera"][0], codes)
        for profile, codes in SERVICE_CODE_PROFILES.items()
    }

    def profile_score(rows: dict[str, OperaRPS]) -> tuple[int, int, int]:
        shared = set(rows) & set(fiscal)
        exact = sum(
            abs(rows[rps].value - fiscal[rps].value) <= Decimal("0.01")
            for rps in shared
        )
        return exact, len(shared), len(rows)

    service_profile, opera = max(
        candidates.items(), key=lambda item: profile_score(item[1])
    )
    fiscal_dates = [item.issue_date for item in fiscal.values() if item.issue_date]
    if not opera:
        raise ValueError(
            "O XML do Opera não contém RPS de serviços reconhecidos para conferência."
        )
    if not fiscal:
        raise ValueError(
            "A planilha Fiscal não contém RPS utilizáveis para conferência."
        )
    if not city:
        raise ValueError(
            "A planilha da Prefeitura não contém RPS utilizáveis para conferência."
        )
    if not fiscal_dates:
        raise ValueError("A planilha Fiscal não possui datas de emissão válidas.")
    fiscal_start, fiscal_end = min(fiscal_dates), max(fiscal_dates)
    results = []
    rps_numbers = sorted(
        set(opera) | set(fiscal) | set(city),
        key=lambda value: (
            not value.isdigit(),
            int(value) if value.isdigit() else value,
        ),
    )
    for rps in rps_numbers:
        op, fi, ci = opera.get(rps), fiscal.get(rps), city.get(rps)
        values = [item.value for item in (op, fi, ci) if item]
        customer = (
            (op.guest if op else "")
            or (fi.customer if fi else "")
            or (ci.customer if ci else "")
        )
        if (
            fi
            and "CANCEL" in normalize(fi.status)
            and ci
            and "CANCEL" in normalize(ci.status)
        ):
            status, detail = (
                "Cancelado",
                "RPS cancelado tanto no Fiscal quanto na Prefeitura.",
            )
        elif fi and "IRREGULAR" in normalize(fi.status):
            status, detail = (
                "Fiscal irregular",
                "O RPS integrou, mas está inválido/irregular e não possui nota válida na Prefeitura.",
            )
        elif op and op.issue_date > fiscal_end and not fi:
            status = "Fora do período do Fiscal"
            detail = (
                f"Não avaliado: este RPS é de {op.issue_date:%d/%m/%Y}, mas o relatório Fiscal "
                f"selecionado contém informações somente até {fiscal_end:%d/%m/%Y}. "
            )
        elif not op:
            status, detail = (
                "Ausente no Opera",
                "RPS localizado nas fontes fiscais, mas não no XML de encerramentos do Opera.",
            )
        elif not fi and not ci:
            status, detail = (
                "Ausente Fiscal/Prefeitura",
                "RPS de serviço encerrado no Opera sem integração Fiscal e sem NFS-e na Prefeitura.",
            )
        elif not fi:
            status, detail = (
                "Ausente no Fiscal",
                "RPS não localizado no módulo Fiscal do CMFlex.",
            )
        elif not ci:
            status, detail = (
                "Ausente na Prefeitura",
                "RPS integrado no Fiscal, mas sem NFS-e correspondente na Prefeitura.",
            )
        elif max(values) - min(values) > Decimal("0.01"):
            status, detail = (
                "Valor divergente",
                "Os valores de serviço não coincidem nas três fontes.",
            )
        else:
            status, detail = (
                "Conciliado",
                "RPS e valor localizados no Opera, Fiscal e Prefeitura.",
            )
        results.append(
            RPSResult(
                rps,
                op.issue_date if op else None,
                op.room if op else "",
                customer,
                op.services if op else "",
                op.value if op else None,
                fi.value if fi else None,
                ci.value if ci else None,
                fi.nfse if fi else "",
                ci.nfse if ci else "",
                fi.status if fi else "Ausente",
                ci.status if ci else "Ausente",
                status,
                detail,
            )
        )
    results.sort(
        key=lambda row: (
            row.reconciled,
            row.incomplete,
            row.cancelled,
            row.opera_date or date.min,
            row.rps,
        )
    )
    return AnalysisResult(
        results,
        fiscal_start,
        fiscal_end,
        tuple(path.name for path in paths),
        service_profile,
    )


def save_excel(result: AnalysisResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Resultado"])
    for label, value in (
        (
            "Período coberto pelo Fiscal",
            f"{result.fiscal_start:%d/%m/%Y} a {result.fiscal_end:%d/%m/%Y}",
        ),
        ("Perfil de serviços identificado", result.service_profile),
        ("RPS analisados", len(result.rows)),
        ("Conciliados", result.reconciled),
        ("Pendências", result.issues),
        ("Fora do período", result.incomplete),
        ("Cancelados", result.cancelled),
    ):
        summary.append([label, value])
    summary.column_dimensions["A"].width = 35
    summary.column_dimensions["B"].width = 42

    detail = workbook.create_sheet("Conferencia_RPS")
    detail.append(
        [
            "RPS",
            "Data Opera",
            "Hóspede/Tomador",
            "Serviços Opera",
            "Valor Opera",
            "Valor Fiscal",
            "Valor Prefeitura",
            "Diferença",
            "NFS-e Fiscal",
            "NFS-e Prefeitura",
            "Situação Fiscal",
            "Situação Prefeitura",
            "Resultado",
            "Explicação",
        ]
    )
    for row in result.rows:
        detail.append(
            [
                row.rps,
                row.opera_date,
                row.customer,
                row.services,
                None if row.opera_value is None else float(row.opera_value),
                None if row.fiscal_value is None else float(row.fiscal_value),
                None if row.city_value is None else float(row.city_value),
                None if row.difference is None else float(row.difference),
                row.fiscal_nfse,
                row.city_nfse,
                row.fiscal_status,
                row.city_status,
                row.status,
                row.detail,
            ]
        )
    for cell in detail["B"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for column in ("E", "F", "G", "H"):
        for cell in detail[column][1:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {
        "A": 13,
        "B": 14,
        "C": 34,
        "D": 40,
        "E": 16,
        "F": 16,
        "G": 18,
        "H": 16,
        "I": 15,
        "J": 18,
        "K": 25,
        "L": 24,
        "M": 28,
        "N": 70,
    }.items():
        detail.column_dimensions[column].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    workbook.save(path)


def save_pdf(result: AnalysisResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    data = [
        ["RPS", "Conciliados", "Pendências", "Fora do período", "Cancelados"],
        [
            str(len(result.rows)),
            str(result.reconciled),
            str(result.issues),
            str(result.incomplete),
            str(result.cancelled),
        ],
    ]
    table = Table(data, colWidths=[35 * mm, 45 * mm, 40 * mm, 50 * mm, 40 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    pending = [row for row in result.rows if not row.reconciled][:35]
    details = [
        ["RPS", "Data", "Hóspede/Tomador", "Opera", "Fiscal", "Prefeitura", "Resultado"]
    ]
    details.extend(
        [
            [
                row.rps,
                row.opera_date.strftime("%d/%m/%Y") if row.opera_date else "—",
                row.customer[:30],
                money(row.opera_value),
                money(row.fiscal_value),
                money(row.city_value),
                row.status,
            ]
            for row in pending
        ]
    )
    detail_table = Table(
        details,
        colWidths=[22 * mm, 25 * mm, 65 * mm, 30 * mm, 30 * mm, 32 * mm, 48 * mm],
        repeatRows=1,
    )
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ]
        )
    )
    story = [
        Paragraph(
            "RPS de serviços prestados — Opera x Fiscal x Prefeitura", styles["Title"]
        ),
        Paragraph(
            f"Perfil identificado: {result.service_profile} • Fiscal de {result.fiscal_start:%d/%m/%Y} a {result.fiscal_end:%d/%m/%Y}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
        table,
        Spacer(1, 6 * mm),
        Paragraph("Registros que exigem atenção", styles["Heading2"]),
        detail_table,
    ]
    SimpleDocTemplate(
        str(path), pagesize=landscape(A4), title="Conferência de RPS"
    ).build(story)
