from __future__ import annotations

import csv
import re
import unicodedata
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

TOLERANCE = Decimal("0.01")


@dataclass(frozen=True)
class CapNote:
    provider: str
    cnpj: str
    number: str
    emission_date: str
    gross: Decimal
    bpm: str
    hotel: str


@dataclass(frozen=True)
class TaxEntry:
    gross: Decimal
    iss: Decimal


@dataclass(frozen=True)
class ResultRow:
    source: str
    provider: str
    cnpj: str
    number: str
    emission_date: str
    gross: Decimal | None
    iss: Decimal | None
    bpm: str
    cap_provider: str
    cap_date: str
    cap_gross: Decimal | None
    cap_iss: Decimal | None
    cap_hotel: str
    status: str

    @property
    def reconciled(self) -> bool:
        return self.status == "Conciliada"

    @property
    def situation(self) -> str:
        if self.reconciled:
            return "Conciliada"
        if "BPM pendente" in self.status:
            return "BPM pendente"
        if "Ausente" in self.status:
            return "Informação ausente"
        return "Divergente"


@dataclass(frozen=True)
class AnalysisResult:
    rows: list[ResultRow]
    external_count: int
    cap_count: int
    matched_count: int
    approved_count: int
    retained_count: int
    cap_retained_count: int
    expected_hotel: str


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^A-Z0-9]",
        "",
        "".join(c for c in text if not unicodedata.combining(c)).upper(),
    )


def cnpj(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def note_number(value: Any) -> str:
    text = str(value or "").strip()
    return (
        str(int(float(text)))
        if re.fullmatch(r"\d+(?:\.0+)?", text)
        else normalize(text)
    )


def decimal_value(value: Any) -> Decimal:
    if value in (None, "", "NULL"):
        return Decimal()
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(str(value).replace(".", "").replace(",", "."))


def date_text(value: Any) -> str:
    if value in (None, "", "NULL"):
        return "—"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip().split()[0]
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return text


def money(value: Decimal) -> str:
    return f"R$ {value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def xlsx_rows(path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    return rows[0], rows[1:]


def xlsx_columns(path: Path) -> set[str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    header = next(sheet.iter_rows(values_only=True), ())
    workbook.close()
    return {normalize(value) for value in header if value not in (None, "")}


def identify_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "external"
    if suffix not in {".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"}:
        return "unknown"
    columns = xlsx_columns(path)
    if {
        "RAZAOSOCIALFORNECEDOR",
        "DOCUMENTOPRINCIPALFORNECEDOR",
        "NUMERO",
        "VALORBRUTO",
        "STATUSBPM",
    }.issubset(columns):
        return "cap"
    if {
        "DOCUMENTOPRINCIPALFORNECEDOR",
        "NUMERODOCUMENTO",
        "VALORBASECALCULO",
        "VALOR",
    }.issubset(columns):
        return "tax"
    if (
        "NUMERONFSE" in columns
        or {"CNPJ", "PRESTADOR", "VALORSERVICOS"}.issubset(columns)
        or {"NUMERO", "PRESTADORDOSERVICO", "DATADEEMISSAO", "VALORDOSERVICO"}.issubset(
            columns
        )
    ):
        return "external"
    return "external" if suffix == ".xls" else "unknown"


def row_dict(header, row):
    return {
        name: row[index] if index < len(row) else None
        for index, name in enumerate(header)
    }


def read_cap(path: Path) -> list[CapNote]:
    header, rows = xlsx_rows(path)
    result = []
    for row in rows:
        data = row_dict(header, row)
        if data.get("Numero") in (None, ""):
            continue
        result.append(
            CapNote(
                str(data.get("RazaoSocialFornecedor") or "").strip(),
                cnpj(data.get("DocumentoPrincipalFornecedor")),
                note_number(data.get("Numero")),
                date_text(data.get("DataEmissao")),
                decimal_value(data.get("ValorBruto")),
                str(data.get("StatusBPM") or "Sem BPM").strip(),
                str(data.get("EmpresaNomeResumido") or "").strip(),
            )
        )
    return result


def read_tax(path: Path) -> dict[tuple[str, str], TaxEntry]:
    header, rows = xlsx_rows(path)
    result = {}
    for row in rows:
        data = row_dict(header, row)
        key = (
            cnpj(data.get("DocumentoPrincipalFornecedor")),
            note_number(data.get("NumeroDocumento")),
        )
        if all(key):
            result[key] = TaxEntry(
                decimal_value(data.get("ValorBaseCalculo")),
                abs(decimal_value(data.get("Valor"))),
            )
    return result


def external_xlsx(path: Path) -> list[dict[str, Any]]:
    header, rows = xlsx_rows(path)
    result = []
    if "Número NFS-e" in header:
        for row in rows:
            data = row_dict(header, row)
            if data.get("Número NFS-e") in (None, ""):
                continue
            retained = str(data.get("Retenção ISSQN") or "").startswith("2")
            result.append(
                {
                    "source": "Portal Nacional",
                    "number": data["Número NFS-e"],
                    "date": data.get("Data Geração"),
                    "cnpj": data.get("CNPJ/CPF Prestador"),
                    "provider": data.get("Nome Prestador"),
                    "gross": data.get("Valor do Serviço (R$)"),
                    "iss": data.get("Valor do ISSQN (R$)") if retained else 0,
                }
            )
    elif "N°" in header:
        for row in rows:
            data = row_dict(header, row)
            if data.get("N°") in (None, ""):
                continue
            result.append(
                {
                    "source": "Prefeitura Caucaia",
                    "number": data["N°"],
                    "date": data.get("DATA"),
                    "cnpj": data.get("CNPJ"),
                    "provider": data.get("PRESTADOR"),
                    "gross": data.get("Valor Serviços"),
                    "iss": data.get("Valor ISS"),
                }
            )
    return result


def external_csv(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="latin1")
    result = []
    for data in csv.DictReader(text.splitlines(), delimiter=";"):
        if not data.get("Nº NFS-e") or data.get("Nº NFS-e") == "Total":
            continue
        result.append(
            {
                "source": "Prefeitura SP",
                "number": data.get("Nº NFS-e"),
                "date": data.get("Data Hora NFE"),
                "cnpj": data.get("CPF/CNPJ do Prestador"),
                "provider": data.get("Razão Social do Prestador"),
                "gross": data.get("Valor dos Serviços"),
                "iss": None,
                "iss_applicable": False,
            }
        )
    return result


def external_html_xls(path: Path) -> list[dict[str, Any]]:
    source_header, source_rows = xlsx_rows(path)
    header = {normalize(value): index for index, value in enumerate(source_header)}

    def value(row, *names: str):
        index = next((header[name] for name in names if name in header), None)
        return row[index] if index is not None and index < len(row) else ""

    required_groups = (
        ("NUMERO", "NUMERONFSE", "NUMERODANOTA", "NOTA"),
        ("PRESTADORDOSERVICO", "PRESTADOR", "RAZAOSOCIAL"),
        ("DATADEEMISSAO", "DATAEMISSAO", "DATA"),
        ("VALORDOSERVICO", "VALORBRUTO", "VALOR"),
    )
    if any(not any(name in header for name in group) for group in required_groups):
        raise ValueError(f"Colunas obrigatórias não encontradas em {path.name}.")

    result = []
    for row in source_rows:
        number = value(row, "NUMERO", "NUMERONFSE", "NUMERODANOTA", "NOTA")
        provider_field = str(
            value(row, "PRESTADORDOSERVICO", "PRESTADOR", "RAZAOSOCIAL") or ""
        )
        match = re.match(r"\s*([\d./-]{14,})\s+-\s+(.+)", provider_field)
        document = match.group(1) if match else value(row, "CNPJ", "CPFCNPJ", "CNPJCPF")
        provider = match.group(2) if match else provider_field
        if not number or not cnpj(document) or not provider:
            continue
        provider = re.sub(r"^\s*[\d./-]{8,}\s+", "", str(provider)).strip()
        result.append(
            {
                "source": "Relatório externo",
                "number": number,
                "date": value(row, "DATADEEMISSAO", "DATAEMISSAO", "DATA"),
                "cnpj": document,
                "provider": provider,
                "gross": value(row, "VALORDOSERVICO", "VALORBRUTO", "VALOR"),
                "iss": value(row, "ISSDEVIDO", "VALORDOISS", "VALORISS", "ISSRETIDO"),
            }
        )
    if not result:
        raise ValueError(
            f"Nenhuma nota com CNPJ e prestador foi encontrada em {path.name}."
        )
    return result


def expected_hotel(paths: list[Path]) -> str:
    aliases = ("TAIBA", "CUMBUCO", "WIND", "MAGNA", "CHARME", "JERI")
    matches = [
        alias for path in paths for alias in aliases if alias in normalize(path.stem)
    ]
    return max(set(matches), key=matches.count) if matches else ""


def bpm_approved(value: Any) -> bool:
    return normalize(value) in {"BMAPROVADO", "BPMAPROVADO"}


def unique_cap_match(
    occurrences: list[dict[str, Any]],
    number: str,
    cap_map: dict[tuple[str, str], CapNote],
    used: set[tuple[str, str]],
) -> tuple[tuple[str, str], CapNote] | None:
    gross_values = [decimal_value(item["gross"]) for item in occurrences]
    candidates = [
        (key, cap)
        for key, cap in cap_map.items()
        if key not in used
        and cap.number == number
        and any(abs(cap.gross - gross) <= TOLERANCE for gross in gross_values)
    ]
    return candidates[0] if len(candidates) == 1 else None


def analyze(paths: list[Path]) -> AnalysisResult:
    if len(paths) < 3:
        raise ValueError(
            "Selecione o CAP, o Alterador ISS e pelo menos uma fonte externa."
        )
    identified = [(path, identify_file(path)) for path in paths]
    cap_paths = [path for path, kind in identified if kind == "cap"]
    tax_paths = [path for path, kind in identified if kind == "tax"]
    external_paths = [path for path, kind in identified if kind == "external"]
    unknown = [path.name for path, kind in identified if kind == "unknown"]
    if unknown:
        raise ValueError(f"Formato não reconhecido: {', '.join(unknown)}.")
    if len(cap_paths) != 1 or len(tax_paths) != 1 or not external_paths:
        raise ValueError(
            "Envie um arquivo CAP, um arquivo de ISS retido e pelo menos uma fonte externa."
        )
    cap_notes, taxes = read_cap(cap_paths[0]), read_tax(tax_paths[0])
    if cap_notes:
        hotel = normalize(
            max(
                {item.hotel for item in cap_notes},
                key=lambda name: sum(row.hotel == name for row in cap_notes),
            )
        )
    else:
        hotel = expected_hotel(paths)
    cap_map = {(item.cnpj, item.number): item for item in cap_notes}
    external = []
    for path in external_paths:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            external.extend(external_csv(path))
        else:
            parsed = external_xlsx(path)
            external.extend(parsed or external_html_xls(path))

    grouped_external = defaultdict(list)
    for raw in external:
        grouped_external[(cnpj(raw["cnpj"]), note_number(raw["number"]))].append(raw)

    result = []
    retained_count = 0
    matched_cap_keys: set[tuple[str, str]] = set()
    for key, occurrences in grouped_external.items():
        raw = occurrences[0]
        sources = " + ".join(sorted({str(item["source"]) for item in occurrences}))
        provider, issued, gross = (
            str(raw["provider"] or "").strip(),
            date_text(raw["date"]),
            decimal_value(raw["gross"]),
        )
        gross_values = [decimal_value(item["gross"]) for item in occurrences]
        iss_sources = [
            item
            for item in occurrences
            if item.get("iss_applicable", True)
            and item.get("iss") not in (None, "")
        ]
        iss = (
            max((decimal_value(item["iss"]) for item in iss_sources), default=Decimal())
            if iss_sources
            else None
        )
        cap = cap_map.get(key)
        cap_key = key
        if cap is None:
            fallback = unique_cap_match(
                occurrences, key[1], cap_map, matched_cap_keys
            )
            if fallback:
                cap_key, cap = fallback
        if cap is not None:
            matched_cap_keys.add(cap_key)
        tax = taxes.get(cap_key) or taxes.get(key)
        issues = []
        if max(gross_values) - min(gross_values) > TOLERANCE:
            issues.append("Valor divergente entre fontes externas")
        bpm = "—"
        if not cap:
            issues.append("Ausente no CAP")
        else:
            sources = f"CAP + {sources}"
            bpm = cap.bpm
            if not bpm_approved(cap.bpm):
                issues.append(f"BPM pendente de aprovação ({cap.bpm})")
            if hotel and cap.hotel and hotel not in normalize(cap.hotel):
                issues.append(f"Hotel divergente ({cap.hotel})")
            if abs(gross - cap.gross) > TOLERANCE:
                issues.append("Valor bruto divergente")
        has_city_source = any(
            str(item["source"]) != "Portal Nacional" for item in occurrences
        )
        if not has_city_source:
            issues.append("Ausente na Prefeitura")
        if iss is not None and iss > 0:
            retained_count += 1
            if not tax:
                issues.append("ISS retido ausente no CAP")
            elif abs(iss - tax.iss) > TOLERANCE:
                issues.append("ISS retido divergente")
        elif tax and tax.iss > 0:
            issues.append("ISS retido ausente na prefeitura")
        result.append(
            ResultRow(
                sources,
                provider,
                key[0],
                key[1],
                issued,
                gross,
                iss,
                bpm,
                cap.provider if cap else "—",
                cap.emission_date if cap else "—",
                cap.gross if cap else None,
                tax.iss if tax else None,
                cap.hotel if cap else "—",
                "Conciliada" if not issues else " • ".join(issues),
            )
        )
    for key, cap in cap_map.items():
        if key in matched_cap_keys:
            continue
        tax = taxes.get(key)
        issues = ["Ausente nas fontes externas"]
        if tax and tax.iss > 0:
            issues.append("ISS retido ausente na prefeitura")
        if not bpm_approved(cap.bpm):
            issues.append(f"BPM pendente de aprovação ({cap.bpm})")
        if hotel and cap.hotel and hotel not in normalize(cap.hotel):
            issues.append(f"Hotel divergente ({cap.hotel})")
        result.append(
            ResultRow(
                "CAP",
                "—",
                cap.cnpj,
                cap.number,
                "—",
                None,
                None,
                cap.bpm,
                cap.provider,
                cap.emission_date,
                cap.gross,
                tax.iss if tax else None,
                cap.hotel,
                " • ".join(issues),
            )
        )
    result.sort(key=lambda row: (row.reconciled, row.source, row.provider, row.number))
    return AnalysisResult(
        result,
        len(grouped_external),
        len(cap_notes),
        sum(row.cap_gross is not None and row.source != "CAP" for row in result),
        sum(bpm_approved(x.bpm) for x in cap_notes),
        retained_count,
        sum(entry.iss > 0 for entry in taxes.values()),
        hotel,
    )


def save_excel(result: AnalysisResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    reconciled = sum(row.reconciled for row in result.rows)
    summary.append(["Indicador", "Resultado"])
    for label, value in (
        ("Hotel esperado", result.expected_hotel),
        ("Notas Prefeitura/Portal únicas", result.external_count),
        ("Notas existentes no arquivo CAP", result.cap_count),
        ("Notas Prefeitura/Portal encontradas no CAP", result.matched_count),
        (
            "Notas Prefeitura/Portal ausentes no CAP",
            result.external_count - result.matched_count,
        ),
        ("Notas existentes somente no CAP", result.cap_count - result.matched_count),
        ("BPM aprovadas no arquivo CAP", result.approved_count),
        ("Notas com ISS retido na prefeitura", result.retained_count),
        ("Notas com ISS retido no CAP", result.cap_retained_count),
        ("Totalmente conciliadas", reconciled),
        ("Com pendências", len(result.rows) - reconciled),
    ):
        summary.append([label, value])
    summary.append([])
    summary.append(["Como usar", "Orientação"])
    summary.append(
        ["Pendências", "Comece por esta aba; motivo e situação aparecem primeiro."]
    )
    summary.append(
        ["Conciliadas", "Notas sem diferença entre as fontes consideradas."]
    )
    summary.append(
        ["Base completa", "Todas as notas e todas as informações disponíveis."]
    )
    for cell in summary[1]:
        cell.style = "Headline 4"
    for cell in summary[14]:
        cell.style = "Headline 4"
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 72

    full_headers = [
        "Situação",
        "Detalhes",
        "Número da nota",
        "CNPJ",
        "Fonte",
        "Prestador Prefeitura/Portal",
        "Prestador CAP",
        "Data externa",
        "Data CAP",
        "Valor externo",
        "Valor CAP",
        "ISS externo",
        "ISS CAP",
        "BPM",
        "Hotel CAP",
    ]
    compact_headers = [
        "Situação",
        "Motivo",
        "Número da nota",
        "CNPJ",
        "Prestador",
        "Fonte",
        "Valor externo",
        "Valor CAP",
        "ISS externo",
        "ISS CAP",
        "BPM",
        "Hotel CAP",
    ]

    def create_detail_sheet(
        title: str,
        rows: list[ResultRow],
        compact: bool,
    ) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(compact_headers if compact else full_headers)
        for row in rows:
            if compact:
                sheet.append(
                    [
                        row.situation,
                        row.status,
                        row.number,
                        row.cnpj,
                        row.provider if row.provider != "—" else row.cap_provider,
                        row.source,
                        float(row.gross) if row.gross is not None else None,
                        float(row.cap_gross) if row.cap_gross is not None else None,
                        float(row.iss) if row.iss is not None else None,
                        float(row.cap_iss) if row.cap_iss is not None else None,
                        row.bpm,
                        row.cap_hotel,
                    ]
                )
            else:
                sheet.append(
                    [
                        row.situation,
                        row.status,
                        row.number,
                        row.cnpj,
                        row.source,
                        row.provider,
                        row.cap_provider,
                        row.emission_date,
                        row.cap_date,
                        float(row.gross) if row.gross is not None else None,
                        float(row.cap_gross) if row.cap_gross is not None else None,
                        float(row.iss) if row.iss is not None else None,
                        float(row.cap_iss) if row.cap_iss is not None else None,
                        row.bpm,
                        row.cap_hotel,
                    ]
                )
        for cell in sheet[1]:
            cell.style = "Headline 4"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        currency_columns = ("G", "H", "I", "J") if compact else ("J", "K", "L", "M")
        for column in currency_columns:
            for cell in sheet[column][1:]:
                cell.number_format = "R$ #,##0.00"
        widths = (
            {
                "A": 20,
                "B": 50,
                "C": 18,
                "D": 18,
                "E": 42,
                "F": 28,
                "G": 18,
                "H": 18,
                "I": 16,
                "J": 16,
                "K": 22,
                "L": 22,
            }
            if compact
            else {
                "A": 20,
                "B": 50,
                "C": 18,
                "D": 18,
                "E": 28,
                "F": 42,
                "G": 42,
                "H": 14,
                "I": 14,
                "J": 18,
                "K": 18,
                "L": 16,
                "M": 16,
                "N": 22,
                "O": 22,
            }
        )
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"
        last_column = "L" if compact else "O"
        sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"

    pending = [row for row in result.rows if not row.reconciled]
    reconciled_rows = [row for row in result.rows if row.reconciled]
    create_detail_sheet("Pendências", pending, True)
    create_detail_sheet("Conciliadas", reconciled_rows, True)
    create_detail_sheet("Base completa", result.rows, False)
    workbook.save(path)


def save_pdf(result: AnalysisResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    total = len(result.rows)
    ok = sum(r.reconciled for r in result.rows)
    data = [
        [
            "Notas analisadas",
            "Encontradas no CAP",
            "BPM aprovadas",
            "Conciliadas",
            "Pendências",
            "ISS Pref./CAP",
        ],
        [
            str(total),
            str(result.matched_count),
            str(result.approved_count),
            str(ok),
            str(total - ok),
            f"{result.retained_count} / {result.cap_retained_count}",
        ],
    ]
    table = Table(data, colWidths=[38 * mm] * 6)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        title="Conferência de Notas de Serviços Tomados",
    )
    doc.build(
        [
            Paragraph("Conferência de Notas de Serviços Tomados", styles["Title"]),
            Spacer(1, 5 * mm),
            table,
        ]
    )
