from __future__ import annotations

import re
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import unquote

import xlrd
from openpyxl import Workbook, load_workbook


def _make_sheet(workbook: Workbook, title: str):
    sheet = workbook.create_sheet(title)
    if not hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions = lambda: None
    return sheet


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag == "table" and self._table is None:
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append(re.sub(r"\s+", " ", "".join(self._cell)).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table is not None:
            if any(self._row):
                self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None


def _workbook_from_rows(tables: list[list[list[Any]]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, rows in enumerate(tables, 1):
        sheet = _make_sheet(workbook, f"Planilha {index}")
        for row in rows:
            sheet.append(list(row))
    if not workbook.worksheets:
        _make_sheet(workbook, "Planilha 1")
    return workbook


def _html_workbook(content: bytes, source: Path) -> Workbook:
    html = ""
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            html = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    parser = _HtmlTableParser()
    parser.feed(html)
    linked_files: list[str] = []
    if not parser.tables:
        references = re.findall(
            r"""(?:href|src)\s*=\s*["']([^"']+\.html?)["']""",
            html,
            flags=re.IGNORECASE,
        )
        for reference in references:
            decoded = unquote(reference).replace("\xa0", " ")
            linked_files.append(decoded)
            candidate = (source.parent / Path(decoded.replace("/", "\\"))).resolve()
            if not candidate.is_file():
                continue
            linked_content = candidate.read_bytes()
            linked_html = ""
            for encoding in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    linked_html = linked_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            parser.feed(linked_html)
    if not parser.tables:
        if linked_files:
            expected = linked_files[0]
            raise ValueError(
                f"{source.name}: este arquivo contém somente o índice de uma página do Excel. "
                f"Os dados deveriam estar em '{expected}', mas essa pasta complementar não foi encontrada. "
                "Exporte ou baixe o relatório como Pasta de Trabalho Excel (.xls ou .xlsx), "
                "e não como Página da Web."
            )
        raise ValueError(f"{source.name}: nenhuma tabela foi encontrada no arquivo .xls.")
    tables = sorted(
        parser.tables,
        key=lambda table: len(table) * max((len(row) for row in table), default=0),
        reverse=True,
    )
    return _workbook_from_rows(tables)


def _xls_workbook(content: bytes, source: Path) -> Workbook:
    try:
        source_workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except xlrd.XLRDError as error:
        raise ValueError(f"{source.name}: arquivo .xls inválido ou corrompido.") from error
    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        for source_sheet in source_workbook.sheets():
            title = source_sheet.name[:31] or "Planilha"
            if title in workbook.sheetnames:
                title = f"{title[:27]}_{len(workbook.sheetnames) + 1}"
            sheet = _make_sheet(workbook, title)
            for row_index in range(source_sheet.nrows):
                values = []
                for column_index in range(source_sheet.ncols):
                    cell = source_sheet.cell(row_index, column_index)
                    value: Any = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(value, source_workbook.datemode)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                        value = int(value)
                    elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                        value = None
                    values.append(value)
                sheet.append(values)
    finally:
        source_workbook.release_resources()
    if not workbook.worksheets:
        raise ValueError(f"{source.name}: nenhuma planilha foi encontrada no arquivo .xls.")
    return workbook


def load_workbook_compatible(
    filename: str | Path,
    *,
    data_only: bool = False,
    read_only: bool = False,
    keep_vba: bool = False,
    **kwargs,
):
    path = Path(filename)
    if path.suffix.lower() != ".xls":
        return load_workbook(
            path,
            data_only=data_only,
            read_only=read_only,
            keep_vba=keep_vba,
            **kwargs,
        )

    content = path.read_bytes()
    if content.startswith(b"PK\x03\x04"):
        return load_workbook(
            BytesIO(content),
            data_only=data_only,
            read_only=read_only,
            keep_vba=keep_vba,
            **kwargs,
        )
    sample = content[:4096].lower()
    if b"<html" in sample or b"<!doctype html" in sample or b"<table" in sample:
        return _html_workbook(content, path)
    return _xls_workbook(content, path)
