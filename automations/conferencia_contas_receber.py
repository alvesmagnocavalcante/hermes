from __future__ import annotations

import warnings
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.common import decimal_value, money as currency
from automations.common import normalize_key as normalize_name
from automations.excel_reader import load_workbook_compatible as load_workbook

TOLERANCE = Decimal("0.01")

CLIENT_GROUPS = (
    ("CVC", ("CVC",)),
    ("DECOLAR / DESPEGAR", ("DECOLAR", "DESPEGAR")),
    ("BESTBUY", ("BESTBUY",)),
    ("BRT", ("BRT",)),
    ("DIVERSA", ("DIVERSA",)),
    ("INTEREP", ("INTEREP",)),
    ("ORINTER", ("ORINTER",)),
    ("PRIMETOUR", ("PRIMETOUR",)),
    ("TREND", ("TREND",)),
    ("VIAGENS PROMO", ("VIAGENSPROMO",)),
    ("VISUAL", ("VISUAL",)),
    ("STONE LINK", ("STONELINK", "LINKSTONE")),
    ("REDE HIPERCARD", ("REDEHIPERCARD", "REDECARDHIPERCARD")),
)


# Modelos dos clientes agrupados, conferências e totais de Contas a Receber.
@dataclass(frozen=True)
class ClientRow:
    client: str
    accounting: Decimal
    financial: Decimal

    @property
    def difference(self) -> Decimal:
        return self.accounting - self.financial

    @property
    def status(self) -> str:
        return "Conciliado" if abs(self.difference) <= TOLERANCE else "Divergente"


@dataclass(frozen=True)
class TotalCheck:
    name: str
    source_value: Decimal
    accounting_value: Decimal

    @property
    def difference(self) -> Decimal:
        return self.source_value - self.accounting_value

    @property
    def status(self) -> str:
        return "Conciliado" if abs(self.difference) <= TOLERANCE else "Divergente"


@dataclass(frozen=True)
class ReceivablesResult:
    clients: list[ClientRow]
    client_accounting_total: Decimal
    client_financial_total: Decimal
    billing: TotalCheck
    commissions: TotalCheck
    hotel: str = "Cumbuco"


def client_group(value: Any) -> tuple[str, str]:
    normalized = normalize_name(value)
    for label, aliases in CLIENT_GROUPS:
        if any(alias in normalized for alias in aliases):
            return normalize_name(label), label
    return normalized, str(value).strip()


# Lê e identifica cada razão contábil e relatório financeiro pelos cabeçalhos.
def load_rows(path: Path) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"{path.name}: planilha vazia.")
        return rows[0], rows[1:]
    finally:
        workbook.close()


def identify_ledger(
    path: Path, header: tuple[Any, ...], rows: list[tuple[Any, ...]]
) -> str | None:
    headers = set(header)
    if not {"DescricaoConta", "Historico"}.issubset(headers):
        return None

    description_i = header.index("DescricaoConta")
    descriptions = {
        normalize_name(row[description_i])
        for row in rows
        if len(row) > description_i and row[description_i] not in (None, "")
    }
    content = "".join(descriptions)
    filename = normalize_name(path.stem)

    if "NOTASAFATURAR" in content:
        return "razao_faturar"
    if "COMISSAO" in content and "CARTAO" in content:
        return "razao_comissao"
    if "NOTASAFATURAR" in filename and "Debito" in headers:
        return "razao_faturar"
    if "COMISSAO" in filename and "Movimento" in headers:
        return "razao_comissao"
    return None


def identify_files(
    paths: list[Path],
) -> dict[str, tuple[tuple[Any, ...], list[tuple[Any, ...]]]]:
    identified: dict[str, tuple[tuple[Any, ...], list[tuple[Any, ...]]]] = {}
    for path in paths:
        header, rows = load_rows(path)
        headers = set(header)
        if {"DescricaoSubconta", "Saldo"}.issubset(headers):
            kind = "balancete"
        elif {"Cliente", "Saldo", "ContaContabilRateio"}.issubset(headers):
            kind = "posicao"
        elif {"Valor", "NumeroDaTransacao", "Status"}.issubset(headers):
            kind = "bordero"
        elif {"NumeroDocumento", "Valor", "IdAgregado"}.issubset(headers):
            kind = "agregados"
        elif ledger_kind := identify_ledger(path, header, rows):
            kind = ledger_kind
        else:
            raise ValueError(
                f"{path.name}: arquivo não reconhecido para esta conferência."
            )
        if kind in identified:
            raise ValueError(f"Dois arquivos foram identificados como {kind}.")
        identified[kind] = (header, rows)
    expected = {
        "balancete",
        "posicao",
        "bordero",
        "razao_faturar",
        "agregados",
        "razao_comissao",
    }
    if set(identified) != expected:
        missing = ", ".join(sorted(expected - set(identified)))
        raise ValueError(
            f"Selecione os seis arquivos da Atividade 8. Ausentes: {missing}."
        )
    return identified


def grouped(
    header: tuple[Any, ...], rows: list[tuple[Any, ...]], name: str, value: str
) -> dict[str, tuple[str, Decimal]]:
    name_i, value_i = header.index(name), header.index(value)
    result: dict[str, tuple[str, Decimal]] = {}
    sums: defaultdict[str, Decimal] = defaultdict(Decimal)
    labels: dict[str, str] = {}
    for row in rows:
        if len(row) <= max(name_i, value_i) or row[name_i] in (None, "", "NULL"):
            continue
        key, label = client_group(row[name_i])
        labels.setdefault(key, label)
        sums[key] += decimal_value(row[value_i])
    for key, value_sum in sums.items():
        result[key] = (labels[key], value_sum)
    return result


def column_total(
    header: tuple[Any, ...],
    rows: list[tuple[Any, ...]],
    column: str,
    absolute: bool = False,
) -> Decimal:
    index = header.index(column)
    total = Decimal()
    for row in rows:
        if len(row) <= index:
            continue
        value = decimal_value(row[index])
        total += abs(value) if absolute else value
    return total


# Consolida clientes e compara Contabilidade e Financeiro por categoria.
def analyze(paths: list[Path], hotel: str = "Cumbuco") -> ReceivablesResult:
    if len(paths) != 6:
        raise ValueError("Selecione exatamente os seis arquivos da Atividade 8.")
    files = identify_files(paths)

    accounting = grouped(*files["balancete"], "DescricaoSubconta", "Saldo")
    financial = grouped(*files["posicao"], "Cliente", "Saldo")
    clients = [
        ClientRow(
            accounting.get(key, financial.get(key))[0],
            accounting.get(key, ("", Decimal()))[1],
            financial.get(key, ("", Decimal()))[1],
        )
        for key in set(accounting) | set(financial)
    ]
    clients.sort(
        key=lambda row: (row.status == "Conciliado", -abs(row.difference), row.client)
    )

    bordero_total = column_total(*files["bordero"], "Valor", absolute=True)
    billing_debit = column_total(*files["razao_faturar"], "Debito")
    aggregate_total = column_total(*files["agregados"], "Valor", absolute=True)
    commission_movement = column_total(
        *files["razao_comissao"], "Movimento", absolute=True
    )
    return ReceivablesResult(
        clients,
        sum((value for _, value in accounting.values()), Decimal()),
        sum((value for _, value in financial.values()), Decimal()),
        TotalCheck("Notas a faturar", bordero_total, billing_debit),
        TotalCheck("Comissões de cartão", aggregate_total, commission_movement),
        hotel,
    )


# Exporta o resultado de Contas a Receber em Excel e PDF.
def save_excel(result: ReceivablesResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Conferência do Contas a Receber"])
    summary.append(["Hotel", result.hotel])
    summary.append([])
    summary.append(["Conferência", "Origem", "Contabilidade", "Diferença", "Status"])
    summary.append(
        [
            "Clientes",
            float(result.client_financial_total),
            float(result.client_accounting_total),
            float(result.client_financial_total - result.client_accounting_total),
            "Conciliado"
            if abs(result.client_financial_total - result.client_accounting_total)
            <= TOLERANCE
            else "Divergente",
        ]
    )
    for check in (result.billing, result.commissions):
        summary.append(
            [
                check.name,
                float(check.source_value),
                float(check.accounting_value),
                float(check.difference),
                check.status,
            ]
        )
    summary["A1"].style = "Title"
    summary["A2"].style = "Headline 4"
    for cell in summary[4]:
        cell.style = "Headline 4"
    for column in ("B", "C", "D"):
        for cell in summary[column][4:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {"A": 28, "B": 20, "C": 20, "D": 18, "E": 16}.items():
        summary.column_dimensions[column].width = width

    details = workbook.create_sheet("Clientes")
    details.append(["Conferência do Contas a Receber"])
    details.append(["Hotel", result.hotel])
    details.append([])
    details.append(
        [
            "Cliente / Subconta",
            "Saldo Contabilidade",
            "Saldo Financeiro",
            "Diferença",
            "Status",
        ]
    )
    for row in result.clients:
        details.append(
            [
                row.client,
                float(row.accounting),
                float(row.financial),
                float(row.difference),
                row.status,
            ]
        )
    details["A1"].style = "Title"
    details["A2"].style = "Headline 4"
    for cell in details[4]:
        cell.style = "Headline 4"
    for column in ("B", "C", "D"):
        for cell in details[column][4:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {"A": 48, "B": 22, "C": 22, "D": 18, "E": 16}.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A5"
    details.auto_filter.ref = f"A4:E{details.max_row}"
    workbook.save(path)


def save_pdf(result: ReceivablesResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Conferência do Contas a Receber",
    )
    client_diff = result.client_financial_total - result.client_accounting_total
    data = [
        ["Conferência", "Origem", "Contabilidade", "Diferença", "Status"],
        [
            "Clientes",
            currency(result.client_financial_total),
            currency(result.client_accounting_total),
            currency(client_diff),
            "Conciliado" if abs(client_diff) <= TOLERANCE else "Divergente",
        ],
    ]
    for check in (result.billing, result.commissions):
        data.append(
            [
                check.name,
                currency(check.source_value),
                currency(check.accounting_value),
                currency(check.difference),
                check.status,
            ]
        )
    table = Table(data, colWidths=[52 * mm, 45 * mm, 45 * mm, 42 * mm, 35 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    document.build(
        [
            Paragraph("Conferência do Contas a Receber", styles["Title"]),
            Spacer(1, 6 * mm),
            table,
        ]
    )
