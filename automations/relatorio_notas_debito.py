from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

HEADERS = [
    "Hotel vendedor",
    "Empresa compradora",
    "Nº da nota",
    "Data de emissão",
    "Item",
    "Valor",
]
STOP_WORDS = (
    "TOTAL LOCACAO",
    "DADOS PARA DEPOSITO",
    "NO CASO DE PAGAMENTO",
    "BANCO",
    "AGENCIA",
    "CONTA",
    "PIX",
    "CNPJ",
    "OBSERVACAO",
    "FORMA DE PAGAMENTO",
    "OPERACAO CC",
)


# Modelo consolidado de cada item encontrado nas notas de débito.
@dataclass(frozen=True)
class ReportRow:
    hotel: str
    comprador: str
    nota: str
    emissao: str
    item: str
    valor: float | int | str

    def values(self) -> list[Any]:
        return [
            self.hotel,
            self.comprador,
            self.nota,
            self.emissao,
            self.item,
            self.valor,
        ]


# Localiza campos variáveis e converte datas, números e valores das planilhas.
def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", "" if value is None else str(value))
    return (
        re.sub(r"\s+", " ", "".join(c for c in text if not unicodedata.combining(c)))
        .strip()
        .upper()
    )


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def format_currency(value: Any) -> str:
    if not isinstance(value, int | float):
        return clean(value)
    formatted = f"R$ {value:,.2f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def find_below(sheet: Worksheet, label: str) -> Any:
    target = normalize(label)
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
        for cell in row:
            if normalize(cell.value) == target:
                return sheet.cell(cell.row + 1, cell.column).value
    return None


def find_header(sheet: Worksheet) -> int | None:
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80)):
        labels = [normalize(cell.value) for cell in row]
        if any(
            label in {"OBJETO", "DESCRICAO", "DESCRICAO ITEM", "LEITURA", "ITEM"}
            for label in labels
        ) and any("V.TOTAL" in label or "VALOR TOTAL" in label for label in labels):
            return row[0].row
    return None


def choose_column(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return next(
        (
            column
            for header, column in headers.items()
            if any(c in header for c in candidates)
        ),
        None,
    )


def seller(sheet: Worksheet, source: Path) -> str:
    for row in sheet.iter_rows(
        min_row=1, max_row=min(sheet.max_row, 8), max_col=min(sheet.max_column, 6)
    ):
        for cell in row:
            text, normalized = clean(cell.value), normalize(cell.value)
            invalid = (
                not text
                or normalized in {"#VALUE!", "#VALOR!"}
                or "@" in text
                or normalized.startswith(
                    ("WWW.", "RUA ", "AV ", "AV:", "AVENIDA ", "RODOVIA ")
                )
                or "CNPJ" in normalized
            )
            if not invalid and not re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", text):
                return text
    return source.stem


def parse_note_number(value: Any, sheet_name: str) -> str:
    text = clean(value)
    match = re.search(
        r"(?:N[°Oº.]?\s*)?([0-9]+(?:[\s./-]*[0-9]+)*)", text, re.IGNORECASE
    )
    if match:
        return re.sub(r"\s+", "", match.group(1))

    sheet_match = re.search(
        r"NF\s*([0-9]+(?:[\s./-]*[0-9]+)*)", sheet_name, re.IGNORECASE
    )
    return re.sub(r"\s+", "", sheet_match.group(1)) if sheet_match else sheet_name


def find_note_value(sheet: Worksheet) -> Any:
    fallback = None
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8)):
        for cell in row:
            value = clean(cell.value)
            if re.match(r"^\s*N[°Oº.]?\s*\d+", value, re.IGNORECASE):
                return value
            if fallback is None and re.search(r"N[°Oº.]?\s*\d+", value, re.IGNORECASE):
                fallback = value
    return fallback


def emission_date(value: Any) -> str:
    if isinstance(value, datetime | date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, int | float):
        try:
            return from_excel(value).strftime("%d/%m/%Y")
        except (TypeError, ValueError):
            pass
    return clean(value)


def money(value: Any) -> float | int | str:
    if value in (None, ""):
        return ""
    if isinstance(value, int | float):
        return round(value, 2)
    text = re.sub(r"[^\d,.-]", "", str(value))
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return value


def extract_sheet(sheet: Worksheet, source: Path) -> list[ReportRow]:
    header_row = find_header(sheet)
    if header_row is None:
        return []
    headers = {
        normalize(cell.value): cell.column
        for cell in sheet[header_row]
        if normalize(cell.value)
    }
    item_col = choose_column(
        headers, ("OBJETO", "DESCRICAO", "DESCRICAO ITEM", "LEITURA", "ITEM")
    )
    value_col = choose_column(
        headers,
        ("V.TOTAL", "VALOR TOTAL", "VALOR TOTAL DO REPASSE", "VALOR TOTAL NOTA"),
    )
    if item_col is None or value_col is None:
        return []

    common = (
        seller(sheet, source),
        clean(find_below(sheet, "Razão social")),
        parse_note_number(find_note_value(sheet), sheet.title),
        emission_date(find_below(sheet, "Data de Emissão")),
    )
    rows: list[ReportRow] = []
    for number in range(header_row + 1, sheet.max_row + 1):
        beginning = " ".join(
            normalize(sheet.cell(number, column).value) for column in range(1, 4)
        )
        if any(word in beginning for word in STOP_WORDS):
            break
        item, value = (
            clean(sheet.cell(number, item_col).value),
            money(sheet.cell(number, value_col).value),
        )
        if item and value not in ("", None, 0, 0.0):
            rows.append(ReportRow(*common, item, value))
    return rows


# Consolida todas as abas válidas dos arquivos selecionados.
def extract(paths: list[Path]) -> list[ReportRow]:
    result: list[ReportRow] = []
    for path in paths:
        workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            for sheet in workbook.worksheets:
                result.extend(extract_sheet(sheet, path))
        finally:
            workbook.close()
    return result


# Exporta o relatório consolidado em Excel e PDF.
def save_excel(rows: list[ReportRow], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row.values())
    for cell in sheet[1]:
        cell.style = "Headline 4"
    for cells in sheet.columns:
        sheet.column_dimensions[cells[0].column_letter].width = min(
            max(max(len(clean(c.value)) for c in cells) + 2, 12), 60
        )
    for cell in sheet["F"][1:]:
        if isinstance(cell.value, int | float):
            cell.number_format = "R$ #,##0.00"
    workbook.save(path)


def save_pdf(rows: list[ReportRow], path: Path) -> None:
    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    body.fontSize = 7
    body.leading = 9
    data = [HEADERS] + [
        [Paragraph(clean(value), body) for value in row.values()[:-1]]
        + [Paragraph(format_currency(row.valor), body)]
        for row in rows
    ]
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    table = Table(
        data,
        repeatRows=1,
        colWidths=[36 * mm, 42 * mm, 22 * mm, 27 * mm, 105 * mm, 28 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F8FAFC")],
                ),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    document.build([table])
