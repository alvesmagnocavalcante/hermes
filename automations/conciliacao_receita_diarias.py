from __future__ import annotations

import re
import unicodedata
import warnings
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook


@dataclass(frozen=True)
class DailyRevenueRow:
    trx_code: str
    description: str
    daily: bool
    average_daily: bool
    transactions: int
    value: Decimal

    @property
    def status(self) -> str:
        return "Com movimento" if self.transactions else "Sem movimento"


@dataclass(frozen=True)
class DailyRevenueResult:
    hotel: str
    rows: list[DailyRevenueRow]
    journal_rows: int

    @property
    def daily_total(self) -> Decimal:
        return sum((row.value for row in self.rows if row.daily), Decimal())

    @property
    def average_daily_total(self) -> Decimal:
        return sum((row.value for row in self.rows if row.average_daily), Decimal())

    @property
    def moved(self) -> int:
        return sum(row.transactions > 0 for row in self.rows)


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^A-Z0-9]",
        "",
        "".join(char for char in text if not unicodedata.combining(char)).upper(),
    )


def trx_code(value: Any) -> str:
    text = str(value or "").strip()
    return (
        str(int(float(text)))
        if re.fullmatch(r"\d+(?:\.0+)?", text)
        else normalize(text)
    )


def decimal_value(value: Any) -> Decimal:
    if value in (None, "", "NULL"):
        return Decimal()
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(str(value).replace(".", "").replace(",", "."))


def money(value: Decimal) -> str:
    return f"R$ {value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def workbook_headers(path: Path) -> list[tuple[str, set[str]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    result = []
    for sheet in workbook.worksheets:
        sheet.reset_dimensions()
        header = next(sheet.iter_rows(values_only=True), ())
        result.append(
            (
                sheet.title,
                {normalize(value) for value in header if value not in (None, "")},
            )
        )
    workbook.close()
    return result


def identify_file(path: Path) -> str:
    headers = workbook_headers(path)
    if any(
        {"TRXCODE", "DIARIA", "DIARIAMEDIA"}.issubset(columns) for _, columns in headers
    ):
        return "codes"
    if any({"TRXCODE", "CASHIERDEBIT"}.issubset(columns) for _, columns in headers):
        return "journal"
    return "unknown"


def read_rules(path: Path, hotel: str) -> dict[str, tuple[str, bool, bool]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = next(
            (
                item
                for item in workbook.worksheets
                if normalize(item.title) == normalize(hotel)
            ),
            None,
        )
        if sheet is None:
            raise ValueError(
                f"O hotel {hotel} não existe na planilha de códigos de transação."
            )
        sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        header = tuple(next(rows))
        indexes = {normalize(value): index for index, value in enumerate(header)}
        required = {"TRXCODE", "DIARIA", "DIARIAMEDIA"}
        if not required.issubset(indexes):
            raise ValueError(
                "A planilha de códigos não contém TRX_CODE, DIÁRIA e DIÁRIA MÉDIA."
            )
        description_index = indexes.get("D3")
        result = {}
        for row in rows:
            code_index = indexes["TRXCODE"]
            if code_index >= len(row) or not trx_code(row[code_index]):
                continue
            daily = (
                indexes["DIARIA"] < len(row)
                and normalize(row[indexes["DIARIA"]]) == "SIM"
            )
            average = (
                indexes["DIARIAMEDIA"] < len(row)
                and normalize(row[indexes["DIARIAMEDIA"]]) == "SIM"
            )
            if not daily and not average:
                continue
            description = (
                str(row[description_index] or "").strip()
                if description_index is not None and description_index < len(row)
                else ""
            )
            result[trx_code(row[code_index])] = description, daily, average
        if not result:
            raise ValueError(
                f"Nenhum TRX_CODE marcado como SIM foi encontrado para {hotel}."
            )
        return result
    finally:
        workbook.close()


def read_journal(path: Path) -> tuple[dict[str, tuple[int, Decimal]], int]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        header = tuple(next(rows))
        indexes = {normalize(value): index for index, value in enumerate(header)}
        if not {"TRXCODE", "CASHIERDEBIT"}.issubset(indexes):
            raise ValueError("O Journal não contém TRX_CODE e CASHIER_DEBIT.")
        grouped: dict[str, list[Any]] = defaultdict(lambda: [0, Decimal()])
        total_rows = 0
        for row in rows:
            code_index, value_index = indexes["TRXCODE"], indexes["CASHIERDEBIT"]
            if code_index >= len(row) or not trx_code(row[code_index]):
                continue
            code = trx_code(row[code_index])
            grouped[code][0] += 1
            grouped[code][1] += decimal_value(
                row[value_index] if value_index < len(row) else None
            )
            total_rows += 1
        return {code: (data[0], data[1]) for code, data in grouped.items()}, total_rows
    finally:
        workbook.close()


def analyze(paths: list[Path], hotel: str) -> DailyRevenueResult:
    if len(paths) != 2:
        raise ValueError("Selecione a planilha de códigos de transação e o Journal.")
    identified = [(path, identify_file(path)) for path in paths]
    codes = [path for path, kind in identified if kind == "codes"]
    journals = [path for path, kind in identified if kind == "journal"]
    unknown = [path.name for path, kind in identified if kind == "unknown"]
    if unknown:
        selected = ", ".join(unknown)
        guidance = (
            "Esta conferência aceita somente a planilha 'Códigos de transação' "
            "e o 'Journal Opera - Receita'."
        )
        if any("BIPDV" in normalize(name) for name in unknown):
            guidance += " O arquivo BI PDV pertence à automação 'Cupons Emitidos x Conta do Hóspede'."
        raise ValueError(
            f"Arquivo incompatível com Receita de Diárias: {selected}. {guidance}"
        )
    if len(codes) != 1 or len(journals) != 1:
        raise ValueError("Envie uma planilha de códigos de transação e um Journal.")
    rules = read_rules(codes[0], hotel)
    journal, journal_rows = read_journal(journals[0])
    rows = [
        DailyRevenueRow(
            code, description, daily, average, *journal.get(code, (0, Decimal()))
        )
        for code, (description, daily, average) in rules.items()
    ]
    rows.sort(key=lambda row: (row.transactions == 0, -abs(row.value), row.trx_code))
    return DailyRevenueResult(hotel, rows, journal_rows)


def save_excel(result: DailyRevenueResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Resultado"])
    for label, value in (
        ("Hotel", result.hotel),
        ("Receita de diárias", float(result.daily_total)),
        ("Receita considerada na diária média", float(result.average_daily_total)),
        ("TRX_CODE considerados", len(result.rows)),
        ("TRX_CODE com movimento", result.moved),
        ("TRX_CODE sem movimento", len(result.rows) - result.moved),
        ("Lançamentos lidos no Journal", result.journal_rows),
    ):
        summary.append([label, value])
    for cell in summary[1]:
        cell.style = "Headline 4"
    for cell in summary["B"][2:4]:
        cell.number_format = "R$ #,##0.00"
    summary.column_dimensions["A"].width = 44
    summary.column_dimensions["B"].width = 24

    details = workbook.create_sheet("Detalhamento")
    details.append(
        [
            "TRX_CODE",
            "Descrição",
            "Diária",
            "Diária média",
            "Lançamentos",
            "CASHIER_DEBIT",
            "Situação",
        ]
    )
    for row in result.rows:
        details.append(
            [
                row.trx_code,
                row.description,
                "SIM" if row.daily else "NÃO",
                "SIM" if row.average_daily else "NÃO",
                row.transactions,
                float(row.value),
                row.status,
            ]
        )
    for cell in details[1]:
        cell.style = "Headline 4"
    for cell in details["F"][1:]:
        cell.number_format = "R$ #,##0.00"
    for column, width in {
        "A": 14,
        "B": 48,
        "C": 14,
        "D": 18,
        "E": 16,
        "F": 20,
        "G": 18,
    }.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    workbook.save(path)


def save_pdf(result: DailyRevenueResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    data = [
        [
            "Hotel",
            "Receita de diárias",
            "Receita diária média",
            "Códigos",
            "Com movimento",
            "Sem movimento",
        ],
        [
            result.hotel,
            money(result.daily_total),
            money(result.average_daily_total),
            str(len(result.rows)),
            str(result.moved),
            str(len(result.rows) - result.moved),
        ],
    ]
    table = Table(
        data, colWidths=[40 * mm, 48 * mm, 48 * mm, 30 * mm, 35 * mm, 35 * mm]
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    document = SimpleDocTemplate(
        str(path), pagesize=landscape(A4), title="Receita de Diárias"
    )
    document.build(
        [
            Paragraph("Conciliação da Receita de Diárias", styles["Title"]),
            Spacer(1, 5 * mm),
            table,
        ]
    )
