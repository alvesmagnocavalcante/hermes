from __future__ import annotations

import re
import unicodedata
import warnings
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as load_openpyxl_workbook

from automations.excel_reader import load_workbook_compatible as load_workbook

DATE_FORMATS = (
    "%d/%m/%Y",
    "%d/%m/%y",
    "%d/%m/%Y %H:%M",
    "%d-%m-%Y",
    "%d-%m-%y",
    "%d-%b-%y",
    "%Y-%m-%d",
)


# Normaliza chaves textuais usadas para relacionar registros entre relatórios.
def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    plain = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]", "", plain.upper())


def identifier(value: Any) -> str:
    text = str(value or "").strip()
    return (
        str(int(float(text)))
        if re.fullmatch(r"\d+(?:\.0+)?", text)
        else normalize_key(text)
    )


# Converte números de diferentes formatos e padroniza a apresentação brasileira.
def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal()
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(str(value).replace(".", "").replace(",", "."))


def nullable_decimal(value: Any) -> Decimal:
    if value in (None, "", "NULL"):
        return Decimal()
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(str(value).replace(".", "").replace(",", "."))


def report_decimal(value: Any) -> Decimal:
    if value in (None, "", "NULL", "-"):
        return Decimal()
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal()


def money(value: Decimal) -> str:
    return f"R$ {value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def optional_money(value: Decimal | None) -> str:
    return "—" if value is None else money(value)


def integer(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def parse_date(value: Any) -> date | None:
    """Converte as representações de data encontradas nos relatórios do HERMES."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip().upper()
    for pattern in DATE_FORMATS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


# Abre a primeira aba e devolve cabeçalho e linhas em um formato comum.
def active_sheet_rows(path: Path) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        sheet.reset_dimensions()
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ValueError(f"{path.name}: planilha vazia.")
    return rows[0], rows[1:]


_INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")


def _source_sheet_title(
    file_stem: str,
    source_title: str,
    multiple_sheets: bool,
    existing: set[str],
) -> str:
    raw_title = f"{file_stem} - {source_title}" if multiple_sheets else file_stem
    base = _INVALID_SHEET_TITLE.sub("-", raw_title).strip(" '") or "Arquivo base"
    candidate = base[:31]
    suffix = 2
    while candidate.casefold() in existing:
        marker = f" ({suffix})"
        candidate = f"{base[: 31 - len(marker)]}{marker}"
        suffix += 1
    existing.add(candidate.casefold())
    return candidate


# Acrescenta ao resultado uma cópia consultável de cada aba dos arquivos-base.
def append_source_workbooks(output: Path, source_paths: Iterable[Path]) -> None:
    paths = [Path(path) for path in source_paths]
    if not paths:
        return

    result = load_openpyxl_workbook(output)
    existing = {title.casefold() for title in result.sheetnames}
    try:
        for path in paths:
            source = load_workbook(path, data_only=True, read_only=True)
            try:
                multiple_sheets = len(source.worksheets) > 1
                for source_sheet in source.worksheets:
                    if hasattr(source_sheet, "reset_dimensions"):
                        source_sheet.reset_dimensions()
                    title = _source_sheet_title(
                        path.stem,
                        source_sheet.title,
                        multiple_sheets,
                        existing,
                    )
                    target = result.create_sheet(title)
                    for row in source_sheet.iter_rows(values_only=True):
                        target.append(list(row))
            finally:
                source.close()
        result.save(output)
    finally:
        result.close()
