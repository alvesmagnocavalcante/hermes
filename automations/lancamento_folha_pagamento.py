from __future__ import annotations

import csv
import os
import re
import unicodedata
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.excel_reader import load_workbook_compatible as load_workbook

SOURCE_LABELS = {
    "summary": "Folha mensal",
    "inss": "INSS mensal",
    "fgts": "FGTS mensal",
    "vacation_receipt": "Recibo de férias",
    "vacation_liquid": "Líquido de férias",
    "vacation_provision": "Provisão de férias",
    "thirteenth_provision": "Provisão de 13º",
}
REQUIRED_SOURCES = set(SOURCE_LABELS)
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ASSETS_DIR = Path(os.environ.get("FLET_ASSETS_DIR", PROJECT_ROOT / "assets")).resolve()
DEFAULT_TEMPLATE = ASSETS_DIR / "folha" / "modelo_folha.xlsm"


@dataclass(frozen=True)
class PostingRow:
    source: str
    source_line: int
    organogram: str
    cost_center: str
    cost_center_name: str
    event: str
    description: str
    debit: str
    credit: str
    value: Decimal
    accounting_origin: int = 6
    employee: str = ""

    @property
    def ready(self) -> bool:
        return bool(self.cost_center and self.debit and self.credit and self.value)

    @property
    def status(self) -> str:
        return "Dados completos" if self.ready else "De/para incompleto"


@dataclass(frozen=True)
class PayrollResult:
    company: str
    period_end: datetime
    earnings: Decimal
    deductions: Decimal
    net_payable: Decimal
    rows: list[PostingRow]
    ignored_rows: int
    excluded_rows: int
    files: dict[str, str]
    warnings: list[str]

    @property
    def total(self) -> Decimal:
        return sum((row.value for row in self.rows), Decimal())

    @property
    def ready(self) -> int:
        return sum(row.ready for row in self.rows)

    @property
    def vacation_employees(self) -> int:
        return len({row.employee for row in self.rows if row.employee})

    @property
    def vacation_entries(self) -> int:
        return sum(bool(row.employee) for row in self.rows)

    @property
    def by_source(self) -> dict[str, list[PostingRow]]:
        grouped: dict[str, list[PostingRow]] = defaultdict(list)
        for row in self.rows:
            grouped[row.source].append(row)
        return dict(grouped)


@dataclass(frozen=True)
class Mappings:
    descriptions: set[str]
    events: dict[str, tuple[str, str]]
    organograms: dict[str, tuple[str, str]]
    organogram_codes: dict[str, tuple[str, str]]
    provisions: dict[str, tuple[str, str]]
    vacations: dict[str, tuple[str, str, str]]
    rate_totals: dict[str, Decimal]
    excluded_events: frozenset[str]


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^A-Z0-9]",
        "",
        "".join(char for char in text if not unicodedata.combining(char)).upper(),
    )


def decimal_value(value: Any) -> Decimal:
    if value in (None, "", "NULL", "-"):
        return Decimal()
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal()


def money(value: Decimal) -> str:
    return f"R$ {value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def read_csv(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    return list(csv.reader(text.splitlines(), delimiter=";"))


def read_tabular(path: Path) -> list[list[str]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        return [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def identify_file(path: Path) -> str:
    if path.suffix.lower() not in {".csv", ".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"}:
        return "unknown"
    if path.suffix.lower() != ".csv":
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(path, data_only=False, read_only=True)
        sheets = {normalize(name) for name in workbook.sheetnames}
        workbook.close()
        if {"DEPARA", "RESULTADO"}.issubset(sheets):
            return "template"
    sample = normalize("\n".join(";".join(row) for row in read_tabular(path)[:80]))
    markers = (
        ("DEMONSTRATIVODEINSS", "inss"),
        ("LIQUIDOSDEFERIAS", "vacation_liquid"),
        ("RECIBODEFERIAS", "vacation_receipt"),
        ("PROVISAODEFERIAS", "vacation_provision"),
        ("PROVISAO13", "thirteenth_provision"),
        ("RELACAODEEVENTOS", "fgts"),
        ("RELACAODECALCULO", "summary"),
    )
    return next((kind for marker, kind in markers if marker in sample), "unknown")


def _account(value: Any) -> str:
    if value in (None, "", "#N/A"):
        return ""
    return (
        str(int(value))
        if isinstance(value, float) and value.is_integer()
        else str(value).strip()
    )


def _cost_center(value: Any) -> str:
    text = _account(value).lstrip("'").strip()
    return text.zfill(4) if text.isdigit() else text


def _organogram_code(value: str) -> str:
    match = re.search(r"(?:ORGRANOGRAMA|ORGANOGRAMA)(\d+(?:\d|\.)*)", normalize(value))
    return match.group(1).strip(".") if match else ""


def read_mappings(path: Path) -> Mappings:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
        formulas = load_workbook(
            path, data_only=False, read_only=False, keep_vba=keep_vba
        )
        values = load_workbook(path, data_only=True, read_only=False, keep_vba=keep_vba)
    try:
        sheet = values["DEPARA"]
        descriptions = {
            normalize(sheet.cell(row, 2).value)
            for row in range(3, sheet.max_row + 1)
            if sheet.cell(row, 2).value
        }
        events: dict[str, tuple[str, str]] = {}
        for row in range(3, sheet.max_row + 1):
            event = sheet.cell(row, 4).value
            if event is not None:
                events[
                    str(int(event)) if isinstance(event, float) else str(event).strip()
                ] = (
                    _account(sheet.cell(row, 7).value),
                    _account(sheet.cell(row, 8).value),
                )

        organograms: dict[str, tuple[str, str]] = {}
        organogram_codes: dict[str, tuple[str, str]] = {}
        for row in range(3, 60):
            source = sheet.cell(row, 12).value
            if not source:
                continue
            mapped = (
                _cost_center(sheet.cell(row, 13).value),
                str(sheet.cell(row, 14).value or "").strip(),
            )
            organograms[normalize(source)] = mapped
            code = _organogram_code(str(source))
            if code:
                organogram_codes[code] = mapped

        provisions = {}
        for row in range(3, 30):
            event = sheet.cell(row, 20).value
            if event:
                provisions[normalize(event)] = (
                    _account(sheet.cell(row, 21).value),
                    _account(sheet.cell(row, 22).value),
                )

        vacations = {}
        for row in range(3, 40):
            code = sheet.cell(row, 26).value
            if code is not None:
                vacations[
                    str(int(code)) if isinstance(code, float) else str(code).strip()
                ] = (
                    str(sheet.cell(row, 27).value or "").strip(),
                    _account(sheet.cell(row, 28).value),
                    _account(sheet.cell(row, 29).value),
                )

        excluded_events = frozenset(
            _account(sheet.cell(row, 16).value)
            for row in range(3, sheet.max_row + 1)
            if sheet.cell(row, 16).value is not None
        )

        rate_events = (
            "REF PLANO ODONTOLOGICO",
            "REF PLANO DE SAUDE",
            "REF INSS MENSAL",
            "REF FGTS MENSAL",
        )
        rate_totals = {
            event: decimal_value(values["C"].cell(2, column).value)
            for event, column in zip(rate_events, range(14, 18))
        }
        return Mappings(
            descriptions,
            events,
            organograms,
            organogram_codes,
            provisions,
            vacations,
            rate_totals,
            excluded_events,
        )
    finally:
        formulas.close()
        values.close()


def lookup_organogram(mappings: Mappings, value: str) -> tuple[str, str]:
    exact = mappings.organograms.get(normalize(value))
    if exact:
        return exact
    return mappings.organogram_codes.get(_organogram_code(value), ("", ""))


def extract_period(rows: list[list[str]]) -> datetime:
    for row in rows:
        text = " ".join(row)
        if "PER" not in normalize(text) and "COMPETENCIA" not in normalize(text):
            continue
        dates = re.findall(r"\d{2}/\d{2}/\d{4}", text)
        if dates:
            return datetime.strptime(dates[-1], "%d/%m/%Y")
        month = re.search(r"(\d{2})/(\d{4})", text)
        if month:
            month_number, year = map(int, month.groups())
            next_month = datetime(year + (month_number == 12), month_number % 12 + 1, 1)
            return datetime.fromtimestamp(next_month.timestamp() - 86400)
    raise ValueError("Não foi possível identificar a competência dos relatórios.")


def extract_payroll_totals(rows: list[list[str]]) -> tuple[Decimal, Decimal, Decimal]:
    totals: dict[str, Decimal] = {}
    for row in rows:
        for index, value in enumerate(row[:-1]):
            marker = normalize(value)
            key = {"PROVVANT": "earnings", "DESCONTOS": "deductions"}.get(marker)
            if marker == "LIQUIDO" and row and normalize(row[0]) == "FGTS":
                key = "net"
            if key is not None:
                amount = decimal_value(row[index + 1])
                if amount:
                    totals[key] = amount
    earnings = totals.get("earnings", Decimal())
    deductions = totals.get("deductions", Decimal())
    net = totals.get("net", earnings - deductions)
    return earnings, deductions, net


def parse_monthly(
    rows: list[list[str]], mappings: Mappings
) -> tuple[list[PostingRow], int]:
    current_organogram = ""
    include = False
    completed = False
    selected_since_total = 0
    ignored = 0
    result: list[PostingRow] = []
    for line_number, row in enumerate(rows, 1):
        first = row[0].strip() if row else ""
        description = row[1].strip() if len(row) > 1 else ""
        marker = normalize(first)
        if marker.startswith(("ORGRANOGRAMA", "ORGANOGRAMA")):
            current_organogram, include, completed, selected_since_total = (
                first,
                False,
                False,
                0,
            )
        if marker.endswith("NORMAL") or marker.startswith("SITUACAO"):
            include = False
        elif marker == "TOTALIRRF":
            if selected_since_total:
                completed, include = True, False
            elif not completed:
                include = True
        is_event = (
            bool(re.fullmatch(r"\d+(?:\.0+)?", first))
            and normalize(description) in mappings.descriptions
        )
        if not is_event:
            continue
        if completed or not include:
            ignored += 1
            continue
        event = str(int(float(first)))
        earnings = decimal_value(row[3] if len(row) > 3 else None)
        deductions = decimal_value(row[4] if len(row) > 4 else None)
        value = earnings or deductions
        if not value:
            continue
        debit, credit = mappings.events.get(event, ("", ""))
        cost_center, cost_center_name = lookup_organogram(mappings, current_organogram)
        result.append(
            PostingRow(
                "Folha mensal",
                line_number,
                current_organogram,
                cost_center,
                cost_center_name,
                event,
                description,
                debit,
                credit,
                value,
            )
        )
        selected_since_total += 1
    return result, ignored


def parse_provision(
    rows: list[list[str]], mappings: Mappings, thirteenth: bool
) -> tuple[list[PostingRow], int]:
    current_organogram = ""
    aggregate_section = False
    result: list[PostingRow] = []
    ignored = 0
    labels = (
        (("Provisão 13°", 1), ("Provisão 13° INSS", 2), ("Provisão 13° FGTS", 3))
        if thirteenth
        else (
            ("Provisão Férias", 1),
            ("1/3 s/ Férias", 2),
            ("Provisão Férias INSS", 3),
            ("Provisão Férias FGTS", 4),
        )
    )
    source = "Provisão de 13º" if thirteenth else "Provisão de férias"
    for line_number, row in enumerate(rows, 1):
        first = row[0].strip() if row else ""
        marker = normalize(first)
        if marker.startswith("ORGANOGRAMA"):
            current_organogram = first
            aggregate_section = False
        elif marker.startswith(("TOTALDAFILIAL", "TOTALDAEMPRESA")):
            aggregate_section = True
        if marker != "PROVISAOMES" or not current_organogram:
            continue
        if aggregate_section:
            ignored += 1
            continue
        cost_center, cost_center_name = lookup_organogram(mappings, current_organogram)
        for description, column in labels:
            value = decimal_value(row[column] if len(row) > column else None)
            if not value:
                continue
            debit, credit = mappings.provisions.get(normalize(description), ("", ""))
            result.append(
                PostingRow(
                    source,
                    line_number,
                    current_organogram,
                    cost_center,
                    cost_center_name,
                    "",
                    description,
                    debit,
                    credit,
                    value,
                )
            )
    return result, ignored


def parse_vacation_employees(
    rows: list[list[str]], mappings: Mappings
) -> tuple[dict[str, tuple[str, str, str]], dict[tuple[str, str, str], Decimal], int]:
    current_organogram = ""
    employees: dict[str, tuple[str, str, str]] = {}
    liquid_totals: dict[tuple[str, str, str], Decimal] = defaultdict(Decimal)
    ignored = 0
    for row in rows:
        first = row[0].strip() if row else ""
        marker = normalize(first)
        if marker.startswith("ORGANOGRAMA"):
            current_organogram = first
        elif marker.startswith("TOTAL"):
            ignored += 1
        elif re.fullmatch(r"\d+", first) and len(row) > 1:
            cost_center, cost_center_name = lookup_organogram(
                mappings, current_organogram
            )
            employees[normalize(row[1])] = (
                cost_center,
                cost_center_name,
                current_organogram,
            )
            liquid_totals[(cost_center, cost_center_name, current_organogram)] += (
                decimal_value(row[8] if len(row) > 8 else None)
            )
    return employees, liquid_totals, ignored


def parse_vacation_receipts(
    rows: list[list[str]],
    employees: dict[str, tuple[str, str, str]],
    mappings: Mappings,
) -> list[PostingRow]:
    employee = ""
    result: list[PostingRow] = []
    for line_number, row in enumerate(rows, 1):
        first = row[0].strip() if row else ""
        if normalize(first).startswith("FUNCIONARIO"):
            employee = next(
                (str(value).strip() for value in row[3:] if str(value).strip()), ""
            )
            continue
        if not employee or not re.fullmatch(r"\d+(?:\.0+)?", first):
            continue
        event = str(int(float(first)))
        if event not in mappings.vacations:
            continue
        description, debit, credit = mappings.vacations[event]
        value = decimal_value(row[4] if len(row) > 4 else None) or decimal_value(
            row[5] if len(row) > 5 else None
        )
        if not value:
            continue
        cost_center, cost_center_name, organogram = employees.get(
            normalize(employee), ("", "", "")
        )
        clean_description = re.sub(r"\s+", " ", description).strip()
        if normalize(clean_description) == "MENSALIDADESINDICALFERIAS":
            clean_description = "Mensalidade Sindical Férias"
        result.append(
            PostingRow(
                "Férias",
                line_number,
                organogram,
                cost_center,
                cost_center_name,
                event,
                f"{employee} {clean_description}",
                debit,
                credit,
                value,
                employee=employee,
            )
        )
    return result


def vacation_receipt_employees(rows: list[list[str]]) -> dict[str, str]:
    employees: dict[str, str] = {}
    for row in rows:
        if not row or not normalize(row[0]).startswith("FUNCIONARIO"):
            continue
        employee = next(
            (str(value).strip() for value in row[3:] if str(value).strip()), ""
        )
        if employee:
            employees[normalize(employee)] = employee
    return employees


def extract_fgts_total(rows: list[list[str]]) -> Decimal:
    totals = [
        decimal_value(row[3])
        for row in rows
        if len(row) > 3 and normalize(row[0]).startswith("QUANTIDADEFUNC")
    ]
    return next((value for value in reversed(totals) if value), Decimal())


def extract_inss_total(rows: list[list[str]]) -> Decimal:
    total_row = next(
        (
            row
            for row in reversed(rows)
            if row and normalize(row[0]).startswith("TOTALEMPRESATIPOCALCULO")
        ),
        [],
    )
    gps_row = next(
        (
            row
            for row in reversed(rows)
            if row and normalize(row[0]).startswith("TOTALGPS")
        ),
        [],
    )
    if len(total_row) <= 7 or len(gps_row) <= 1:
        return Decimal()
    return (
        decimal_value(gps_row[1])
        - decimal_value(total_row[4])
        - decimal_value(total_row[7])
    )


def allocate(total: Decimal, weights: dict[str, Decimal]) -> dict[str, Decimal]:
    active = [(key, value) for key, value in sorted(weights.items()) if value > 0]
    denominator = sum((value for _, value in active), Decimal())
    if not total or not denominator:
        return {}
    result: dict[str, Decimal] = {}
    assigned = Decimal()
    for index, (key, weight) in enumerate(active):
        value = (
            total - assigned
            if index == len(active) - 1
            else (total * weight / denominator).quantize(Decimal("0.01"), ROUND_HALF_UP)
        )
        result[key] = value
        assigned += value
    return result


def build_rates(
    monthly: list[PostingRow],
    mappings: Mappings,
    inss_total: Decimal,
    fgts_total: Decimal,
    period_end: datetime,
) -> tuple[list[PostingRow], list[str]]:
    weights: dict[str, Decimal] = defaultdict(Decimal)
    names: dict[str, str] = {}
    for row in monthly:
        if row.cost_center:
            weights[row.cost_center] += row.value
            names[row.cost_center] = row.cost_center_name
    totals = dict(mappings.rate_totals)
    validation: list[str] = []
    for event, source_total in (
        ("REF INSS MENSAL", inss_total),
        ("REF FGTS MENSAL", fgts_total),
    ):
        model_total = totals.get(event, Decimal())
        if source_total:
            totals[event] = source_total
            if model_total and abs(model_total - source_total) > Decimal("0.01"):
                validation.append(
                    f"{event}: modelo {money(model_total)}; relatório atual {money(source_total)}. Foi usado o relatório."
                )
    result: list[PostingRow] = []
    for event, total in totals.items():
        debit, credit = mappings.provisions.get(normalize(event), ("", ""))
        for cost_center, value in allocate(total, weights).items():
            result.append(
                PostingRow(
                    "Rateios mensais",
                    0,
                    "Rateio proporcional da folha",
                    cost_center,
                    names.get(cost_center, ""),
                    "",
                    event,
                    debit,
                    credit,
                    value,
                )
            )
    return result, validation


def _default_template() -> Path | None:
    return DEFAULT_TEMPLATE if DEFAULT_TEMPLATE.exists() else None


def analyze(paths: list[Path]) -> PayrollResult:
    identified = [(path, identify_file(path)) for path in paths]
    unknown = [path.name for path, kind in identified if kind == "unknown"]
    if unknown:
        raise ValueError(
            f"Arquivo não reconhecido pelo conteúdo: {', '.join(unknown)}."
        )
    grouped: dict[str, list[Path]] = defaultdict(list)
    for path, kind in identified:
        grouped[kind].append(path)
    duplicated = [
        SOURCE_LABELS.get(kind, kind)
        for kind, items in grouped.items()
        if kind != "template" and len(items) > 1
    ]
    if duplicated:
        raise ValueError(f"Há mais de um arquivo para: {', '.join(duplicated)}.")
    missing = [SOURCE_LABELS[kind] for kind in SOURCE_LABELS if not grouped.get(kind)]
    if missing:
        raise ValueError(
            f"Faltam {len(missing)} relatório(s): {', '.join(sorted(missing))}."
        )
    templates = grouped.get("template", [])
    template = templates[0] if len(templates) == 1 else _default_template()
    if len(templates) > 1:
        raise ValueError("Selecione no máximo uma planilha modelo Excel.")
    if not template:
        raise ValueError("A planilha modelo da atividade 2 não foi encontrada.")

    sources = {kind: read_tabular(grouped[kind][0]) for kind in REQUIRED_SOURCES}
    mappings = read_mappings(template)
    period_end = extract_period(sources["summary"])
    earnings, deductions, net_payable = extract_payroll_totals(sources["summary"])
    company = next(
        (
            row[0].strip()
            for row in sources["summary"]
            if row and "HOTEL" in normalize(row[0])
        ),
        "Empresa não identificada",
    )

    monthly, monthly_ignored = parse_monthly(sources["summary"], mappings)
    vacation_provision, vacation_ignored = parse_provision(
        sources["vacation_provision"], mappings, False
    )
    thirteenth, thirteenth_ignored = parse_provision(
        sources["thirteenth_provision"], mappings, True
    )
    employees, liquid_totals, liquid_ignored = parse_vacation_employees(
        sources["vacation_liquid"], mappings
    )
    vacations = parse_vacation_receipts(
        sources["vacation_receipt"], employees, mappings
    )
    receipt_employees = vacation_receipt_employees(sources["vacation_receipt"])
    generated_employees = {normalize(row.employee) for row in vacations if row.employee}
    missing_employees = [
        name
        for key, name in receipt_employees.items()
        if key not in generated_employees
    ]
    if missing_employees:
        raise ValueError(
            "Não foi possível gerar os lançamentos de férias de: "
            f"{', '.join(missing_employees)}. Verifique o de/para dos eventos."
        )
    incomplete_vacations = sorted(
        {row.employee for row in vacations if row.employee and not row.ready}
    )
    if incomplete_vacations:
        raise ValueError(
            "Funcionário(s) de férias sem conta ou centro de custo mapeado: "
            f"{', '.join(incomplete_vacations)}."
        )
    individual_vacations = {row for row in vacations if row.employee}
    for (cost_center, cost_center_name, organogram), value in liquid_totals.items():
        if value:
            vacations.append(
                PostingRow(
                    "Férias",
                    10**9,
                    organogram,
                    cost_center,
                    cost_center_name,
                    "",
                    "REF BAIXA FÉRIAS",
                    "201010103",
                    "101020103",
                    value.quantize(Decimal("0.01")),
                )
            )
    rates, validation = build_rates(
        monthly,
        mappings,
        extract_inss_total(sources["inss"]),
        extract_fgts_total(sources["fgts"]),
        period_end,
    )
    rows = monthly + vacations + vacation_provision + thirteenth + rates
    # A lista "Desconsiderar" elimina do resumo mensal os eventos que já são
    # detalhados pelos relatórios específicos. Aplicá-la novamente ao recibo
    # removeria todos os funcionários em férias da exportação.
    excluded_rows = sum(
        row.source == "Folha mensal" and row.event in mappings.excluded_events
        for row in rows
    )
    rows = [
        row
        for row in rows
        if not (row.source == "Folha mensal" and row.event in mappings.excluded_events)
    ]
    exported_individual_vacations = {row for row in rows if row.employee}
    if individual_vacations != exported_individual_vacations:
        raise ValueError(
            "A validação interna detectou lançamento individual de férias ausente no resultado."
        )
    if not rows:
        raise ValueError(
            "Nenhum lançamento foi encontrado nos relatórios selecionados."
        )
    rows.sort(
        key=lambda item: (
            item.source,
            not item.ready,
            item.source_line if item.source == "Férias" else 0,
            item.cost_center,
            item.description,
            item.source_line,
        )
    )
    files = {SOURCE_LABELS[kind]: grouped[kind][0].name for kind in REQUIRED_SOURCES}
    return PayrollResult(
        company,
        period_end,
        earnings,
        deductions,
        net_payable,
        rows,
        monthly_ignored + vacation_ignored + thirteenth_ignored + liquid_ignored,
        excluded_rows,
        files,
        validation,
    )


def _write_import_sheet(
    workbook: Workbook, title: str, rows: list[PostingRow], posting_date: datetime
) -> None:
    sheet = workbook.create_sheet(title)
    for row in rows:
        if not row.ready:
            continue
        sheet.append(_posting_values(row, posting_date, excel=True))
    sheet.column_dimensions["B"].width = 13
    sheet.column_dimensions["J"].width = 58
    for cell in sheet["B"]:
        cell.number_format = "dd/mm/yyyy"
    for column in ("F", "I"):
        for cell in sheet[column]:
            cell.number_format = "@"
    for cell in sheet["M"]:
        cell.number_format = "#,##0.00"


def _csv_money(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def _posting_values(
    row: PostingRow, posting_date: datetime, excel: bool = False
) -> list[Any]:
    cost_center = _cost_center(row.cost_center)
    return [
        "A",
        posting_date if excel else posting_date.strftime("%d/%m/%Y"),
        "",
        int(row.debit),
        "",
        cost_center,
        int(row.credit),
        "",
        cost_center,
        row.description,
        "",
        row.accounting_origin,
        float(row.value) if excel else _csv_money(row.value),
    ]


def save_csv(
    result: PayrollResult,
    path: Path,
    posting_date: datetime | None = None,
    source: str | None = None,
) -> None:
    posting_date = posting_date or result.period_end
    exported_rows = [
        row
        for row in result.rows
        if row.ready and (source is None or row.source == source)
    ]
    with path.open("w", encoding="cp1252", newline="") as file:
        # O Excel/CMFlex em ambiente pt-BR usa ponto e vírgula como separador
        # de campos e vírgula como separador decimal.
        writer = csv.writer(
            file, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n"
        )
        writer.writerows(_posting_values(row, posting_date) for row in exported_rows)

    saved_rows = read_csv(path)
    if len(saved_rows) != len(exported_rows):
        raise ValueError("O CSV gerado não preservou todos os lançamentos.")
    for line, (saved, source_row) in enumerate(zip(saved_rows, exported_rows), 1):
        expected = _cost_center(source_row.cost_center)
        if len(saved) < 9 or saved[5] != expected or saved[8] != expected:
            raise ValueError(
                f"O centro de custo não foi preservado nas colunas F e I do CSV, linha {line}."
            )


def save_excel(
    result: PayrollResult, path: Path, posting_date: datetime | None = None
) -> None:
    posting_date = posting_date or result.period_end
    workbook = Workbook()
    workbook.remove(workbook.active)
    names = {
        "Folha mensal": "Folha_Mensal",
        "Férias": "Ferias",
        "Provisão de férias": "Provisao_Ferias",
        "Provisão de 13º": "Provisao_13",
        "Rateios mensais": "Rateios_Mensais",
    }
    for source, title in names.items():
        _write_import_sheet(
            workbook, title, result.by_source.get(source, []), posting_date
        )

    summary = workbook.create_sheet("Resumo", 0)
    summary.append(["Indicador", "Resultado"])
    indicators = [
        ("Empresa", result.company),
        ("Competência", result.period_end.strftime("%m/%Y")),
        ("Relatórios reconhecidos", len(result.files)),
        ("Lançamentos gerados", len(result.rows)),
        ("Lançamentos com dados completos", result.ready),
        ("De/para incompleto", len(result.rows) - result.ready),
        ("Proventos da folha", float(result.earnings)),
        ("Descontos da folha", float(result.deductions)),
        ("Líquido a pagar", float(result.net_payable)),
        ("Totalizadores/duplicadores excluídos", result.ignored_rows),
        ("Eventos da lista Desconsiderar excluídos", result.excluded_rows),
        ("Funcionários em férias incluídos", result.vacation_employees),
        ("Lançamentos individuais de férias", result.vacation_entries),
        ("Valor total dos lançamentos", float(result.total)),
    ]
    for label, value in indicators:
        summary.append([label, value])
    summary.append([])
    summary.append(["Origem", "Arquivo", "Lançamentos", "Valor"])
    for source, filename in sorted(result.files.items()):
        label = SOURCE_LABELS.get(
            next((key for key, value in SOURCE_LABELS.items() if value == source), ""),
            source,
        )
        posting_source = "Férias" if source == "Recibo de férias" else source
        rows = result.by_source.get(posting_source, [])
        summary.append(
            [
                label,
                filename,
                len(rows),
                float(sum((row.value for row in rows), Decimal())),
            ]
        )
    if result.warnings:
        summary.append([])
        summary.append(["Validações"])
        for warning in result.warnings:
            summary.append([warning])
    for cell in summary[1]:
        cell.style = "Headline 4"
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 75
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 22
    for row in summary.iter_rows():
        for cell in row:
            if cell.column == 4 and isinstance(cell.value, float):
                cell.number_format = "R$ #,##0.00"

    details = workbook.create_sheet("Detalhamento")
    details.append(
        [
            "Fonte",
            "Linha origem",
            "Organograma",
            "Centro de custo",
            "Nome do CC",
            "Evento",
            "Descrição",
            "Débito",
            "Crédito",
            "Valor",
            "Situação",
        ]
    )
    for row in result.rows:
        details.append(
            [
                row.source,
                row.source_line or "Rateio",
                row.organogram,
                row.cost_center,
                row.cost_center_name,
                row.event,
                row.description,
                row.debit,
                row.credit,
                float(row.value),
                row.status,
            ]
        )
    for cell in details[1]:
        cell.style = "Headline 4"
    for cell in details["J"][1:]:
        cell.number_format = "R$ #,##0.00"
    for column, width in {
        "A": 23,
        "B": 14,
        "C": 42,
        "D": 16,
        "E": 25,
        "F": 12,
        "G": 58,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 23,
    }.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    workbook.save(path)


def save_pdf(result: PayrollResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    data = [["Saída", "Lançamentos", "Valor"]]
    for source, rows in result.by_source.items():
        data.append(
            [source, str(len(rows)), money(sum((row.value for row in rows), Decimal()))]
        )
    data.append(["TOTAL", str(len(result.rows)), money(result.total)])
    table = Table(data, colWidths=[80 * mm, 45 * mm, 55 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -2),
                    [colors.white, colors.HexColor("#EEF3F8")],
                ),
            ]
        )
    )
    vacation_groups: dict[str, list[PostingRow]] = defaultdict(list)
    for row in result.rows:
        if row.employee:
            vacation_groups[row.employee].append(row)
    vacation_data = [
        ["Funcionário em férias", "Centro de custo", "Lançamentos", "Valor"]
    ]
    for employee, rows in sorted(vacation_groups.items()):
        vacation_data.append(
            [
                employee,
                rows[0].cost_center,
                str(len(rows)),
                money(sum((row.value for row in rows), Decimal())),
            ]
        )
    vacation_table = Table(
        vacation_data, colWidths=[100 * mm, 35 * mm, 30 * mm, 35 * mm]
    )
    vacation_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EEF3F8")],
                ),
            ]
        )
    )
    story = [
        Paragraph("Lançamento da Folha de Pagamento", styles["Title"]),
        Spacer(1, 3 * mm),
        Paragraph(
            f"{result.company} — competência {result.period_end:%m/%Y}",
            styles["BodyText"],
        ),
        Paragraph(f"Líquido a pagar: {money(result.net_payable)}", styles["Heading2"]),
        Paragraph(
            f"Exclusões aplicadas: {result.ignored_rows} totalizadores e "
            f"{result.excluded_rows} eventos duplicados do resumo mensal.",
            styles["BodyText"],
        ),
        Paragraph(
            f"Férias: {result.vacation_employees} funcionário(s) e "
            f"{result.vacation_entries} lançamento(s) individualizado(s) incluídos.",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
        table,
    ]
    if vacation_groups:
        story.extend(
            [
                Spacer(1, 7 * mm),
                Paragraph("Funcionários em férias incluídos", styles["Heading2"]),
                Spacer(1, 2 * mm),
                vacation_table,
            ]
        )
    document = SimpleDocTemplate(
        str(path), pagesize=landscape(A4), title="Lançamento da Folha de Pagamento"
    )
    document.build(story)
