from __future__ import annotations

import re
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from automations.common import normalize_key as normalize, optional_money as money
from automations.excel_reader import load_workbook_compatible as load_workbook


# Modelos das fontes e do resultado da conciliação de cupons por hóspede.
@dataclass(frozen=True)
class CouponResult:
    company: str
    pdv: str
    issue_date: date
    posting_date: date | None
    document: str
    account: str
    room: str
    guest: str
    document_type: str
    pdv_value: Decimal
    journal_value: Decimal | None
    status: str
    detail: str

    @property
    def difference(self) -> Decimal | None:
        return (
            None if self.journal_value is None else self.pdv_value - self.journal_value
        )

    @property
    def reconciled(self) -> bool:
        return self.status.startswith("Conciliado")

    @property
    def incomplete_period(self) -> bool:
        return self.status == "Journal não cobre a data"


@dataclass(frozen=True)
class ReconciliationResult:
    company: str
    mapping: str
    journal_start: date
    journal_end: date
    coupons: list[CouponResult]
    files: tuple[str, str, str]

    @property
    def reconciled(self) -> int:
        return sum(item.reconciled for item in self.coupons)

    @property
    def incomplete(self) -> int:
        return sum(item.incomplete_period for item in self.coupons)

    @property
    def issues(self) -> int:
        return len(self.coupons) - self.reconciled - self.incomplete

    @property
    def total_pdv(self) -> Decimal:
        return sum((item.pdv_value for item in self.coupons), Decimal())


@dataclass
class _Coupon:
    company: str
    pdv: str
    issue_date: date
    document: str
    account: str
    room: str
    guest: str
    document_type: str
    value: Decimal = Decimal()


@dataclass(frozen=True)
class _JournalRow:
    code: str
    check: str
    posting_date: date
    value: Decimal
    room: str


def decimal_value(value: Any) -> Decimal:
    if value in (None, "", "-"):
        return Decimal()
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal()


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _header_map(values: tuple[Any, ...]) -> dict[str, int]:
    result = {}
    for index, value in enumerate(values):
        key = normalize(value)
        if key and key not in result:
            result[key] = index
    return result


# Identifica e lê PDV, Journal e tabela de relacionamento de contas.
def identify_file(path: Path) -> str:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls", ".xltx", ".xltm"}:
        return "unknown"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        first_headers = _header_map(
            tuple(next(workbook.active.iter_rows(values_only=True)))
        )
        keys = set(first_headers)
        if {
            "EMPRESA",
            "PDV",
            "VALOR",
            "DATADEEMISSAO",
            "NODODOCUMENTO",
            "CONTA",
            "DADOSHOSPEDE",
        }.issubset(keys):
            return "pdv"
        if {
            "TRXCODE",
            "REFERENCE",
            "CASHIERDEBIT",
            "BUSINESSFORMATDATE",
            "ROOM",
        }.issubset(keys):
            return "journal"
        if workbook.sheetnames and all(
            normalize(next(sheet.iter_rows(values_only=True))[0]) == "TRXCODE"
            for sheet in workbook.worksheets
        ):
            return "mapping"
        return "unknown"
    finally:
        workbook.close()


def _read_pdv(path: Path) -> dict[tuple[str, str, date, str], _Coupon]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = _header_map(tuple(next(rows)))

        def value(row, name):
            index = headers[name]
            return row[index] if index < len(row) else None

        coupons: dict[tuple[str, str, date, str], _Coupon] = {}
        for row in rows:
            company = str(value(row, "EMPRESA") or "").strip()
            account = re.sub(r"\D", "", str(value(row, "CONTA") or ""))
            issued = parse_date(value(row, "DATADEEMISSAO"))
            document = str(value(row, "NODODOCUMENTO") or "").strip()
            if (
                not company
                or not account
                or not issued
                or not document
                or normalize(company) in {"TOTAL", "FILTROSAPLICADOS"}
            ):
                continue
            guest_data = str(value(row, "DADOSHOSPEDE") or "").strip()
            parts = [part.strip() for part in guest_data.split("/")]
            key = (company, account, issued, document)
            if key not in coupons:
                coupons[key] = _Coupon(
                    company,
                    str(value(row, "PDV") or "").strip(),
                    issued,
                    document,
                    account,
                    parts[0].zfill(4)
                    if parts and parts[0].isdigit()
                    else (parts[0] if parts else ""),
                    parts[1] if len(parts) > 1 else "",
                    str(value(row, "TIPODEDOCUMENTO") or "").strip(),
                )
            coupons[key].value += decimal_value(value(row, "VALOR"))
        for coupon in coupons.values():
            coupon.value = coupon.value.quantize(Decimal("0.01"))
        return coupons
    finally:
        workbook.close()


def _read_journal(path: Path) -> list[_JournalRow]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = _header_map(tuple(next(rows)))

        def value(row, name):
            index = headers[name]
            return row[index] if index < len(row) else None

        result = []
        for row in rows:
            reference = str(value(row, "REFERENCE") or "")
            match = re.search(r"CHECK#\s*(\d+)", reference, re.IGNORECASE)
            posting_date = parse_date(value(row, "BUSINESSFORMATDATE"))
            if not match or not posting_date:
                continue
            code_value = value(row, "TRXCODE")
            code = (
                str(int(code_value))
                if isinstance(code_value, float) and code_value.is_integer()
                else str(code_value or "").strip()
            )
            result.append(
                _JournalRow(
                    code,
                    match.group(1),
                    posting_date,
                    decimal_value(value(row, "CASHIERDEBIT")),
                    str(value(row, "ROOM") or "").strip().zfill(4),
                )
            )
        return result
    finally:
        workbook.close()


def _read_mappings(path: Path) -> dict[str, set[str]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        result = {}
        for sheet in workbook.worksheets:
            codes = set()
            for (value,) in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if value is not None:
                    codes.add(
                        str(int(value))
                        if isinstance(value, float) and value.is_integer()
                        else str(value).strip()
                    )
            result[sheet.title] = codes
        return result
    finally:
        workbook.close()


def _match_account(check: str, accounts: set[str]) -> str | None:
    if check in accounts:
        return check
    matches = [account for account in accounts if check.endswith(account)]
    return max(matches, key=len) if matches else None


def _mapping_matches_company(mapping_name: str, company: str) -> bool:
    """Indica se a aba do de/para pertence explicitamente à empresa do BI/PDV."""
    mapping = normalize(mapping_name)
    company_name = normalize(company)
    if mapping and mapping in company_name:
        return True
    return (mapping == "WIND" and "CUMBUCO" in company_name) or (
        mapping == "CUMBUCO" and "WIND" in company_name
    )


# Relaciona cupom, check e conta do hóspede e classifica cada ocorrência.
def analyze(paths: list[Path]) -> ReconciliationResult:
    if len(paths) != 3:
        raise ValueError(
            "Selecione o BI/PDV, o Journal e o arquivo de de/para dos TRX_CODE."
        )
    identified = [(path, identify_file(path)) for path in paths]
    unknown = [path.name for path, kind in identified if kind == "unknown"]
    if unknown:
        raise ValueError(
            f"Arquivo não reconhecido pelo conteúdo: {', '.join(unknown)}."
        )
    grouped = {
        kind: [path for path, current in identified if current == kind]
        for kind in ("pdv", "journal", "mapping")
    }
    invalid = [kind for kind, items in grouped.items() if len(items) != 1]
    if invalid:
        raise ValueError(
            "Envie exatamente um BI/PDV, um Journal e um arquivo de de/para."
        )

    pdv = _read_pdv(grouped["pdv"][0])
    journal = _read_journal(grouped["journal"][0])
    mappings = _read_mappings(grouped["mapping"][0])
    if not pdv or not journal or not mappings:
        raise ValueError("Um dos arquivos não contém registros utilizáveis.")

    companies: dict[str, set[str]] = defaultdict(set)
    for coupon in pdv.values():
        companies[coupon.company].add(coupon.account)
    best: tuple[int, int, str, str] | None = None
    for mapping_name, codes in mappings.items():
        checks = [row.check for row in journal if row.code in codes]
        for company, accounts in companies.items():
            matched = len(
                {_match_account(check, accounts) for check in checks} - {None}
            )
            candidate = (
                int(_mapping_matches_company(mapping_name, company)),
                matched,
                mapping_name,
                company,
            )
            if best is None or candidate > best:
                best = candidate
    if not best or best[1] == 0:
        raise ValueError(
            "Não foi encontrada relação entre as contas do BI/PDV e os CHECK# do Journal."
        )
    _, _, mapping_name, company = best
    accounts = companies[company]
    selected_journal = [row for row in journal if row.code in mappings[mapping_name]]
    journal_start = min(row.posting_date for row in selected_journal)
    journal_end = max(row.posting_date for row in selected_journal)
    postings: dict[str, dict[date, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    for row in selected_journal:
        account = _match_account(row.check, accounts)
        if account:
            postings[account][row.posting_date] += row.value

    results = []
    for coupon in (item for item in pdv.values() if item.company == company):
        by_date = postings.get(coupon.account, {})
        posting_date = coupon.issue_date if coupon.issue_date in by_date else None
        detail = "Conta e valor localizados no Journal."
        if posting_date is None and by_date:
            exact = [
                day
                for day, value in by_date.items()
                if abs(value - coupon.value) <= Decimal("0.01")
            ]
            if exact:
                posting_date = min(
                    exact, key=lambda day: (abs((day - coupon.issue_date).days), day)
                )
        if posting_date is not None:
            journal_value = by_date[posting_date].quantize(Decimal("0.01"))
            difference = coupon.value - journal_value
            if journal_value == 0 and coupon.value != 0:
                status, detail = (
                    "Não cobrado",
                    "O CHECK# existe, mas o valor líquido lançado no Journal é zero.",
                )
            elif abs(difference) > Decimal("0.01"):
                status, detail = (
                    "Valor divergente",
                    f"Diferença de {money(difference)} entre o cupom e a conta.",
                )
            elif posting_date != coupon.issue_date:
                status = "Conciliado - data diferente"
                detail = f"Cobrado no Journal em {posting_date:%d/%m/%Y}."
            else:
                status = "Conciliado"
        elif coupon.issue_date < journal_start or coupon.issue_date > journal_end:
            journal_value = None
            status = "Journal não cobre a data"
            detail = f"O Journal selecionado cobre {journal_start:%d/%m/%Y} a {journal_end:%d/%m/%Y}."
        elif by_date:
            nearest = min(by_date, key=lambda day: abs((day - coupon.issue_date).days))
            journal_value = by_date[nearest].quantize(Decimal("0.01"))
            posting_date = nearest
            status = "Lançado em outra data"
            detail = f"CHECK# localizado em {nearest:%d/%m/%Y}, mas sem correspondência segura de valor."
        else:
            journal_value = None
            status = "Ausente na conta"
            detail = "Cupom emitido no BI/PDV sem CHECK# correspondente no Journal."
        results.append(
            CouponResult(
                company,
                coupon.pdv,
                coupon.issue_date,
                posting_date,
                coupon.document,
                coupon.account,
                coupon.room,
                coupon.guest,
                coupon.document_type,
                coupon.value,
                journal_value,
                status,
                detail,
            )
        )
    results.sort(
        key=lambda item: (
            item.reconciled,
            item.incomplete_period,
            item.issue_date,
            item.account,
        )
    )
    return ReconciliationResult(
        company,
        mapping_name,
        journal_start,
        journal_end,
        results,
        tuple(path.name for path in paths),
    )


# Exporta detalhes e resumo da conciliação em Excel e PDF.
def save_excel(result: ReconciliationResult, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Resultado"])
    for label, value in (
        ("Hotel", result.company),
        ("De/para identificado", result.mapping),
        (
            "Período do Journal",
            f"{result.journal_start:%d/%m/%Y} a {result.journal_end:%d/%m/%Y}",
        ),
        ("Cupons analisados", len(result.coupons)),
        ("Conciliados", result.reconciled),
        ("Pendências", result.issues),
        ("Fora do período", result.incomplete),
        ("Valor total dos cupons", float(result.total_pdv)),
    ):
        summary.append([label, value])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 42
    summary["B9"].number_format = "R$ #,##0.00"

    detail = workbook.create_sheet("Conferencia")
    detail.append(
        [
            "Hotel",
            "Data cupom",
            "Data Journal",
            "PDV",
            "Cupom",
            "Conta/CHECK",
            "Quarto",
            "Hóspede",
            "Valor cupom",
            "Valor Journal",
            "Diferença",
            "Resultado",
            "Explicação",
        ]
    )
    for item in result.coupons:
        detail.append(
            [
                item.company,
                item.issue_date,
                item.posting_date,
                item.pdv,
                item.document,
                item.account,
                item.room,
                item.guest,
                float(item.pdv_value),
                None if item.journal_value is None else float(item.journal_value),
                None if item.difference is None else float(item.difference),
                item.status,
                item.detail,
            ]
        )
    for column in ("B", "C"):
        for cell in detail[column][1:]:
            cell.number_format = "dd/mm/yyyy"
    for column in ("I", "J", "K"):
        for cell in detail[column][1:]:
            cell.number_format = "R$ #,##0.00"
    for column, width in {
        "A": 22,
        "B": 14,
        "C": 14,
        "D": 20,
        "E": 12,
        "F": 16,
        "G": 10,
        "H": 32,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 28,
        "M": 65,
    }.items():
        detail.column_dimensions[column].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    workbook.save(path)


def save_pdf(result: ReconciliationResult, path: Path) -> None:
    styles = getSampleStyleSheet()
    data = [
        ["Cupons", "Conciliados", "Pendências", "Fora do período", "Valor total"],
        [
            str(len(result.coupons)),
            str(result.reconciled),
            str(result.issues),
            str(result.incomplete),
            money(result.total_pdv),
        ],
    ]
    table = Table(data, colWidths=[35 * mm, 40 * mm, 35 * mm, 45 * mm, 50 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24588A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    pending = [item for item in result.coupons if not item.reconciled][:30]
    pending_data = [["Data", "Cupom", "Conta", "Hóspede", "Valor", "Resultado"]]
    pending_data.extend(
        [
            [
                item.issue_date.strftime("%d/%m/%Y"),
                item.document,
                item.account,
                item.guest[:28],
                money(item.pdv_value),
                item.status,
            ]
            for item in pending
        ]
    )
    pending_table = Table(
        pending_data,
        colWidths=[25 * mm, 25 * mm, 30 * mm, 65 * mm, 30 * mm, 55 * mm],
        repeatRows=1,
    )
    pending_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ]
        )
    )
    story = [
        Paragraph("Cupons emitidos x conta do hóspede", styles["Title"]),
        Paragraph(
            f"{result.company} — Journal de {result.journal_start:%d/%m/%Y} a {result.journal_end:%d/%m/%Y}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
        table,
    ]
    if pending:
        story.extend(
            [
                Spacer(1, 6 * mm),
                Paragraph("Cupons que exigem atenção", styles["Heading2"]),
                pending_table,
            ]
        )
    SimpleDocTemplate(
        str(path), pagesize=landscape(A4), title="Conferência de cupons"
    ).build(story)
