from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from automations.common import append_source_workbooks
from hermes_ui.registry import SOURCE_TABS_AUTOMATIONS, SPECS


class SourceWorkbookTabsTest(TestCase):
    def test_appends_every_source_sheet_after_analysis_tabs(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "resultado.xlsx"
            first = root / "Razao [Contabilidade].xlsx"
            second_dir = root / "outra"
            second_dir.mkdir()
            second = second_dir / first.name

            result = Workbook()
            result.active.title = "Análise"
            result.create_sheet("Divergências")
            result.save(output)

            source = Workbook()
            source.active.title = "Dados"
            source.active.append(["Documento", "Valor"])
            source.active.append([101, 25.5])
            source.create_sheet("Resumo").append(["Total", 25.5])
            source.save(first)

            duplicate = Workbook()
            duplicate.active.append(["Outro arquivo"])
            duplicate.save(second)

            append_source_workbooks(output, (first, second))

            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames[:2], ["Análise", "Divergências"])
            self.assertEqual(len(workbook.sheetnames), 5)
            self.assertEqual(workbook.worksheets[2]["A2"].value, 101)
            self.assertEqual(workbook.worksheets[3]["B1"].value, 25.5)
            self.assertEqual(workbook.worksheets[4]["A1"].value, "Outro arquivo")
            self.assertEqual(len({name.casefold() for name in workbook.sheetnames}), 5)
            self.assertTrue(all(len(name) <= 31 for name in workbook.sheetnames))
            workbook.close()

    def test_scope_is_limited_to_requested_automations(self):
        self.assertEqual(
            SOURCE_TABS_AUTOMATIONS,
            {"receita", "cupons", "receber", "pagar", "custos"},
        )

    def test_every_hotel_automation_includes_hotel_in_filename(self):
        result = type("Result", (), {"hotel": "Charme"})()
        hotel_specs = [spec for spec in SPECS if spec.hotel_option]

        self.assertEqual(
            {spec.key for spec in hotel_specs},
            {"receita", "diarias", "cupons", "receber", "pagar"},
        )
        for spec in hotel_specs:
            with self.subTest(spec=spec.key):
                self.assertEqual(
                    spec.output_filename(result, "Cumbuco", "Excel"),
                    f"{spec.key}_charme_resultado.xlsx",
                )
