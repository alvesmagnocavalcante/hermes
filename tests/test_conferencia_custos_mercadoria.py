from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from automations.conferencia_custos_mercadoria import (
    analyze,
    entry_postings,
    export_excel,
    final_balances,
)


def create_workbook(path: Path, header: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class MerchandiseCostsTest(TestCase):
    def test_entry_comparison_ignores_cost_and_stock_movements(self):
        data = (
            ("DescricaoConta", "Debito", "Historico"),
            [
                (
                    "Alimentos",
                    Decimal("100"),
                    "Lançamento Nota Fiscal Eletrônica de Mercadoria (Terceiros)",
                ),
                (
                    "Alimentos",
                    Decimal("70"),
                    "Saída por transferência - integração de custo",
                ),
                ("Alimentos", Decimal("5"), "Ajuste de saldo negativo"),
            ],
        )

        self.assertEqual(entry_postings(data), {"Alimentos": Decimal("100")})

    def test_reads_balance_from_row_below_account_without_movement(self):
        data = (
            ("DescricaoConta", "SaldoAtual"),
            [("Alimentos", None), (None, Decimal("125.50"))],
        )

        self.assertEqual(final_balances(data), {"Alimentos": Decimal("125.50")})

    def test_analyze_keeps_final_balance_for_account_without_movement(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            documents = root / "DOCUMENTOSLANCADOS.xlsx"
            entries = root / "RAZAOANALITICOESTOQUEAB.xlsx"
            inventory = root / "INVENTARIOFISICO.xlsx"
            stock = root / "RAZAOANALITICOESTOQUES.xlsx"

            create_workbook(documents, ["DESCRICAOTDESEMB", "VALORLANÇADO"], [])
            create_workbook(entries, ["DescricaoConta", "Debito", "Historico"], [])
            create_workbook(
                inventory,
                ["GrupoCodigo", "SaldoValor"],
                [["01", Decimal("125.50")]],
            )
            create_workbook(
                stock,
                ["DescricaoConta", "SaldoAtual"],
                [["Alimentos", None], [None, Decimal("125.50")]],
            )

            rows = analyze([documents, entries, inventory, stock])
            alimentos = next(
                row
                for row in rows
                if row.analysis == "Saldo final" and row.account == "Alimentos"
            )

            self.assertEqual(alimentos.accounting, Decimal("125.50"))
            self.assertEqual(alimentos.status, "Conciliado")

            result_path = root / "resultado.xlsx"
            export_excel(rows, result_path)
            result = load_workbook(result_path, read_only=True)
            self.assertEqual(result.sheetnames, ["Entradas", "Saldo final"])
            self.assertEqual(
                [cell.value for cell in result["Entradas"][1]],
                ["Conta", "CAP", "Contabilidade", "Diferença", "Status"],
            )
            self.assertEqual(
                [cell.value for cell in result["Saldo final"][1]],
                ["Conta", "Inventário", "Contabilidade", "Diferença", "Status"],
            )
            result.close()
