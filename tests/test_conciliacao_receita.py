from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from automations.conciliacao_receita import reconcile, save_excel_result


class ReconcileRevenueTest(TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.accounting = root / "contabilidade.xlsx"
        self.journal = root / "journal.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Documento", "DataLancamento", "Movimento"])
        sheet.append(["NULL", "01/01/0001", 0])
        sheet.append(["1", "04/07/2026", -100])
        sheet.append(["2", "13/07/2026", -200])
        sheet.append(["3", "20/07/2026", -50])
        workbook.save(self.accounting)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "TRX_NO",
                "BUSINESS_FORMAT_DATE",
                "CASHIER_DEBIT",
                "TRX_DESC",
            ]
        )
        sheet.append(["1", "04/07/26", 100, "Diária"])
        sheet.append(["2", "04/07/26", 780, "Multa Avaria"])
        sheet.append(["3", "13/07/26", 200, "Diária"])
        sheet.append(["4", "13/07/26", 10, "Avaria"])
        sheet.append(["5", "13/07/26", 30, "Ajuste Taxa de Turismo"])
        sheet.append(["6", "13/07/26", 20, "Taxa de Turismo"])
        sheet.append(["7", "13/07/26", 40, "A Faturar"])
        sheet.append(["8", "13/07/26", 15, "Turismo"])
        sheet.append(["9", "20-07-26", 50, "Diária"])
        sheet.append(["10", "20-07-26", 25, "Taxa Extra"])
        workbook.save(self.journal)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_magna_ignores_only_tourism_fees_by_business_date(self):
        result = reconcile([self.accounting, self.journal], "Magna")

        self.assertEqual(result.cmflex_total, Decimal("350"))
        self.assertEqual(result.opera_total, Decimal("1220"))
        self.assertEqual(len(result.rows), 31)
        self.assertEqual(result.rows[0].business_date, date(2026, 7, 1))
        self.assertEqual(result.rows[-1].business_date, date(2026, 7, 31))
        row = next(
            row for row in result.rows if row.business_date == date(2026, 7, 13)
        )
        self.assertEqual(row.difference, Decimal("-65"))
        row = next(
            row for row in result.rows if row.business_date == date(2026, 7, 4)
        )
        self.assertEqual(row.difference, Decimal("-780"))
        self.assertEqual(row.identification, "Multa Avaria (Journal): R$ 780,00")
        row = next(
            row for row in result.rows if row.business_date == date(2026, 7, 20)
        )
        self.assertEqual(row.identification, "Taxa Extra (Journal): R$ 25,00")

    def test_other_hotels_keep_tourism_fees_and_damage(self):
        result = reconcile([self.accounting, self.journal], "Cumbuco")

        self.assertEqual(result.hotel, "Cumbuco")
        self.assertEqual(result.opera_total, Decimal("1270"))
        row = next(
            row for row in result.rows if row.business_date == date(2026, 7, 13)
        )
        self.assertEqual(row.difference, Decimal("-115"))

    def test_export_identifies_selected_hotel(self):
        result = reconcile([self.accounting, self.journal], "Charme")
        output = Path(self.temp_dir.name) / "resultado.xlsx"

        save_excel_result(result, output)

        workbook = load_workbook(output, data_only=True)
        self.assertEqual(workbook["Resumo"]["B2"].value, "Charme")
        self.assertEqual(workbook["Conciliação"]["B1"].value, "Charme")
        workbook.close()
