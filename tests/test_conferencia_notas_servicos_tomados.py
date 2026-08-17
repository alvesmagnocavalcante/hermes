from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from automations.conferencia_notas_servicos_tomados import (
    analyze,
    decimal_value,
    external_csv,
    external_html_xls,
    identify_file,
    save_excel,
)


def create_cap(
    path: Path,
    cnpj: str,
    number: str,
    gross: Decimal,
    bpm: str = "BPM APROVADO",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "RazaoSocialFornecedor",
            "DocumentoPrincipalFornecedor",
            "Numero",
            "DataEmissao",
            "ValorBruto",
            "StatusBPM",
            "EmpresaNomeResumido",
        ]
    )
    sheet.append(
        [
            "PRESTADOR NO CAP",
            cnpj,
            number,
            "01/07/2026",
            gross,
            bpm,
            "CARMEL TAÍBA",
        ]
    )
    workbook.save(path)


def create_tax(
    path: Path,
    cnpj: str | None = None,
    number: str | None = None,
    gross: Decimal = Decimal(),
    iss: Decimal = Decimal(),
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "DocumentoPrincipalFornecedor",
            "NumeroDocumento",
            "ValorBaseCalculo",
            "Valor",
        ]
    )
    if cnpj and number:
        sheet.append([cnpj, number, gross, iss])
    workbook.save(path)


def create_city(
    path: Path,
    cnpj: str,
    number: str,
    gross: Decimal,
    iss: Decimal,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        ["N°", "DATA", "CNPJ", "PRESTADOR", "Valor Serviços", "Valor ISS"]
    )
    sheet.append(
        [number, "02/07/2026", cnpj, "PRESTADOR PREFEITURA", gross, iss]
    )
    workbook.save(path)


class ServiceNotesTest(TestCase):
    def test_matches_portal_number_with_national_prefix_to_cap(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap.xlsx"
            tax = root / "tax.xlsx"
            portal = root / "portal.xlsx"
            document = "39906832000107"

            create_cap(cap, document, "119115", Decimal("50"))
            create_tax(tax)
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Número NFS-e",
                    "Retenção ISSQN",
                    "Data Geração",
                    "CNPJ/CPF Prestador",
                    "Nome Prestador",
                    "Valor do Serviço (R$)",
                    "Valor do ISSQN (R$)",
                ]
            )
            sheet.append(
                [
                    "2600000119115",
                    "1 - Não retido",
                    "01/08/2026",
                    document,
                    "IFT SERVICOS DE TELECOMUNICACOES LTDA",
                    50,
                    0,
                ]
            )
            workbook.save(portal)

            rows = analyze([cap, tax, portal]).rows

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].number, "119115")
            self.assertEqual(rows[0].cap_gross, Decimal("50"))
            self.assertIn("CAP + Portal Nacional", rows[0].source)

    def test_recognizes_municipal_services_taken_layout(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "NF-141111-072026-ServicosTomados.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Tipo Doc.",
                    "Número",
                    "Data",
                    "Valor dos Serviços",
                    "ISS Retido",
                    "Valor do ISS",
                    "CPF/CNPJ Prestador",
                    "Razão Social/Nome do Prestador",
                ]
            )
            sheet.append(
                [
                    "NFS-e",
                    "395",
                    "27/07/2026",
                    18634.18,
                    "Sim",
                    719.28,
                    "37694753000154",
                    "A C CONSTRUCOES CTS CRUZ LTDA",
                ]
            )
            workbook.save(path)

            self.assertEqual(identify_file(path), "external")
            row = external_html_xls(path)[0]
            self.assertEqual(row["source"], "Prefeitura")
            self.assertEqual(row["number"], "395")
            self.assertEqual(row["cnpj"], "37694753000154")
            self.assertEqual(row["provider"], "A C CONSTRUCOES CTS CRUZ LTDA")
            self.assertEqual(row["gross"], 18634.18)
            self.assertEqual(row["iss"], 719.28)

    def test_non_numeric_iss_markers_are_zero(self):
        self.assertEqual(decimal_value("Não Retido"), Decimal())
        self.assertEqual(decimal_value("—"), Decimal())
        self.assertEqual(decimal_value("-"), Decimal())
        self.assertEqual(decimal_value("R$ 1.234,56"), Decimal("1234.56"))

    def test_uses_portal_retained_iss_and_ignores_name_differences(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap (TAIBA).xlsx"
            tax = root / "iss (TAIBA).xlsx"
            portal = root / "portal (TAIBA).xlsx"
            city = root / "prefeitura (TAIBA).xlsx"
            document = "12345678000190"

            create_cap(cap, document, "10", Decimal("100"))
            create_tax(
                tax, document, "10", Decimal("100"), Decimal("5")
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Número NFS-e",
                    "Retenção ISSQN",
                    "Data Geração",
                    "CNPJ/CPF Prestador",
                    "Nome Prestador",
                    "Valor do Serviço (R$)",
                    "Valor do ISSQN (R$)",
                ]
            )
            sheet.append(
                [
                    "10",
                    "2 - Retido",
                    "02/07/2026",
                    document,
                    "NOME DIFERENTE NO PORTAL",
                    100,
                    5,
                ]
            )
            workbook.save(portal)
            create_city(city, document, "10", Decimal("100"), Decimal("5"))

            result = analyze([cap, tax, portal, city])
            row = result.rows[0]

            self.assertTrue(row.reconciled)
            self.assertEqual(result.expected_hotel, "CARMELTAIBA")
            self.assertEqual(row.iss, Decimal("5"))
            self.assertEqual(row.provider, "NOME DIFERENTE NO PORTAL")
            self.assertEqual(row.cap_provider, "PRESTADOR NO CAP")
            self.assertNotIn("Data divergente", row.status)
            self.assertNotIn("Razão social divergente", row.status)

    def test_spreadsheet_from_sao_paulo_does_not_compare_iss(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "prefeitura-sp.csv"
            path.write_text(
                "Nº NFS-e;Data Hora NFE;CPF/CNPJ do Prestador;"
                "Razão Social do Prestador;Valor dos Serviços;"
                "ISS Retido;ISS devido\n"
                "10;02/07/2026;12345678000190;"
                "PRESTADOR;100,00;S;5,00\n",
                encoding="latin1",
            )

            row = external_csv(path)[0]

            self.assertIsNone(row["iss"])
            self.assertFalse(row["iss_applicable"])

    def test_matches_unique_cap_note_by_number_and_value_when_cnpj_differs(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap (TAIBA).xlsx"
            tax = root / "iss (TAIBA).xlsx"
            city = root / "prefeitura (TAIBA).xlsx"

            create_cap(cap, "11111111000111", "2932", Decimal("3242"))
            create_tax(tax)
            create_city(
                city,
                "22222222000122",
                "2932",
                Decimal("3242"),
                Decimal(),
            )

            result = analyze([cap, tax, city])

            self.assertEqual(len(result.rows), 1)
            self.assertEqual(result.matched_count, 1)
            self.assertTrue(result.rows[0].reconciled)

            output = root / "resultado.xlsx"
            save_excel(result, output)
            exported = load_workbook(output)
            try:
                self.assertEqual(
                    exported.sheetnames,
                    ["Resumo", "Pendências", "Conciliadas", "Base completa"],
                )
                self.assertEqual(
                    exported["Base completa"]["A1"].value, "Situação"
                )
                self.assertEqual(
                    exported["Base completa"]["B1"].value, "Detalhes"
                )
                self.assertEqual(exported["Pendências"].freeze_panes, "A2")
                self.assertEqual(exported["Conciliadas"].freeze_panes, "A2")
                self.assertEqual(exported["Base completa"].freeze_panes, "A2")
                self.assertEqual(len(exported["Pendências"].tables), 0)
                self.assertEqual(len(exported["Conciliadas"].tables), 0)
                self.assertEqual(len(exported["Base completa"].tables), 0)
            finally:
                exported.close()

    def test_bpm_in_approval_is_pending_not_unposted(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap (TAIBA).xlsx"
            tax = root / "iss (TAIBA).xlsx"
            city = root / "prefeitura (TAIBA).xlsx"

            create_cap(
                cap,
                "11111111000111",
                "159",
                Decimal("11500"),
                "BPM em Aprovação",
            )
            create_tax(tax)
            create_city(
                city,
                "11111111000111",
                "159",
                Decimal("11500"),
                Decimal(),
            )

            row = analyze([cap, tax, city]).rows[0]

            self.assertEqual(row.situation, "BPM pendente")
            self.assertIn("BPM pendente de aprovação", row.status)
            self.assertNotIn("Não escriturada", row.status)

    def test_portal_only_note_reports_missing_city_source(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap (TAIBA).xlsx"
            tax = root / "iss (TAIBA).xlsx"
            portal = root / "portal (TAIBA).xlsx"
            document = "12345678000190"

            create_cap(cap, document, "1", Decimal("17500"))
            create_tax(tax)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Número NFS-e",
                    "Retenção ISSQN",
                    "Data Geração",
                    "CNPJ/CPF Prestador",
                    "Nome Prestador",
                    "Valor do Serviço (R$)",
                    "Valor do ISSQN (R$)",
                ]
            )
            sheet.append(
                [
                    "1",
                    "1 - Não retido",
                    "20/07/2026",
                    document,
                    "PRESTADOR",
                    17500,
                    0,
                ]
            )
            workbook.save(portal)

            row = analyze([cap, tax, portal]).rows[0]

            self.assertIn("Ausente na Prefeitura", row.status)

    def test_prefeitura_provider_without_spaces_around_hyphen_is_matched(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            cap = root / "cap (TAIBA).xlsx"
            tax = root / "iss (TAIBA).xlsx"
            portal = root / "portal (TAIBA).xlsx"
            city = root / "PREF (TAIBA).xlsx"
            document = "60245303000104"

            create_cap(cap, document, "145", Decimal("2250"))
            create_tax(tax)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Número NFS-e",
                    "Retenção ISSQN",
                    "Data Geração",
                    "CNPJ/CPF Prestador",
                    "Nome Prestador",
                    "Valor do Serviço (R$)",
                    "Valor do ISSQN (R$)",
                ]
            )
            sheet.append(
                [
                    "145",
                    "1 - Não retido",
                    "06/07/2026",
                    document,
                    "F DA ROCHA MARTINS TRANSPORTES LTDA",
                    2250,
                    0,
                ]
            )
            workbook.save(portal)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Número",
                    "Prestador do Serviço",
                    "Data de Emissão",
                    "Valor do Serviço",
                    "ISS Devido",
                ]
            )
            sheet.append(
                [
                    "145",
                    "60.245.303/0001-04-F DA ROCHA MARTINS TRANSPORTES LTDA",
                    "01/07/2026",
                    2250,
                    0,
                ]
            )
            workbook.save(city)

            row = analyze([cap, tax, portal, city]).rows[0]

            self.assertTrue(row.reconciled)
            self.assertIn("Prefeitura", row.source)
            self.assertNotIn("Ausente na Prefeitura", row.status)

    def test_recognizes_charme_prefeitura_layout(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "PREF OFC.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Numero",
                    "Data",
                    "Mes competencia",
                    "Ano competencia",
                    "Forma entrada",
                    "Doc prestador",
                    "Nome prestador",
                    "Item servico",
                    "Valor faturado",
                    "Base de calculo",
                    "Aliquota",
                    "Valor iss",
                    "Iss retido",
                ]
            )
            sheet.append(
                [
                    "388",
                    "31/07/2026",
                    7,
                    2026,
                    "Manual",
                    "21002293000116",
                    "WEMBLEY VIAGENS E TURISMO LTDA",
                    "902",
                    "R$ 9.803,47",
                    "R$ 9.803,47",
                    0,
                    "R$ 51,92",
                    "NÃO",
                ]
            )
            workbook.save(path)

            rows = external_html_xls(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["number"], "388")
            self.assertEqual(rows[0]["cnpj"], "21002293000116")
            self.assertEqual(rows[0]["provider"], "WEMBLEY VIAGENS E TURISMO LTDA")
            self.assertEqual(rows[0]["gross"], "R$ 9.803,47")
            self.assertEqual(rows[0]["iss"], 0)
            self.assertEqual(rows[0]["source"], "Prefeitura")
