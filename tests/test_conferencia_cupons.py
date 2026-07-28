from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from automations.conferencia_cupons import read_file, reconcile, save_excel


class CouponReconciliationTest(TestCase):
    def test_aprovado_c_is_included_as_simphony(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "simphony.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave da NF", "Valor Total NF", "Status", "Data"])
            sheet.append(["CHAVE-1", 150, "Aprovado (C)", "01/07/2026"])
            sheet.append(["CHAVE-2", 80, "Cancelado", "01/07/2026"])
            workbook.save(path)

            source, values, cancelled = read_file(path)

            self.assertEqual(source, "Simphony")
            self.assertEqual(values["CHAVE-1"].value, Decimal("150"))
            self.assertEqual(values["CHAVE-2"].status, "Cancelado")
            self.assertEqual(cancelled, 1)

    def test_cancelled_simphony_coupon_is_not_reported_as_absent(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            simphony = root / "simphony.xlsx"
            fiscal = root / "fiscal.xlsx"
            sefaz = root / "sefaz.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave da NF", "Valor Total NF", "Status", "Data"])
            sheet.append(["CHAVE-CANCELADA", 80, "Cancelado", "01/07/2026"])
            workbook.save(simphony)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave", "ValorContabil", "DataDocumento", "Cancelado"])
            sheet.append(["CHAVE-CANCELADA", 0, "01/07/2026", False])
            workbook.save(fiscal)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave de acesso", "Valor R$", "Data de emissão"])
            sheet.append(["CHAVE-CANCELADA", 0, "01/07/2026"])
            workbook.save(sefaz)

            result = reconcile([simphony, fiscal, sefaz], "Magna")

            self.assertEqual(result.rows[0].status, "Conciliado: cancelado")
            self.assertEqual(result.rows[0].simphony_status, "Cancelado")
            self.assertEqual(result.rows[0].difference, Decimal())
            self.assertEqual(result.simphony_total, Decimal())
            self.assertEqual(result.hotel, "Magna")

            output = root / "resultado.xlsx"
            save_excel(result, output)
            exported = load_workbook(output, read_only=True)
            try:
                self.assertEqual(exported["Resumo"]["B2"].value, "Magna")
                self.assertEqual(
                    exported["Conciliação"]["H1"].value, "Status Simphony"
                )
                self.assertEqual(
                    exported["Conciliação"]["H2"].value, "Cancelado"
                )
            finally:
                exported.close()

    def test_cancelled_coupon_with_downstream_value_is_divergent(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            simphony = root / "simphony.xlsx"
            fiscal = root / "fiscal.xlsx"
            sefaz = root / "sefaz.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave da NF", "Valor Total NF", "Status", "Data"])
            sheet.append(["CHAVE-CANCELADA", 80, "Cancelado", "01/07/2026"])
            workbook.save(simphony)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave", "ValorContabil", "DataDocumento", "Cancelado"])
            sheet.append(["CHAVE-CANCELADA", 80, "01/07/2026", False])
            workbook.save(fiscal)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave de acesso", "Valor R$", "Data de emissão"])
            sheet.append(["CHAVE-CANCELADA", 80, "01/07/2026"])
            workbook.save(sefaz)

            result = reconcile([simphony, fiscal, sefaz])

            self.assertEqual(result.rows[0].status, "Divergente: cancelamento")
            self.assertEqual(result.rows[0].difference, Decimal("80"))

    def test_cancelled_coupon_with_empty_fiscal_and_sefaz_is_reconciled(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            simphony = root / "simphony.xlsx"
            fiscal = root / "fiscal.xlsx"
            sefaz = root / "sefaz.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Chave da NF", "Valor Total NF", "Status", "Data"])
            sheet.append(["CHAVE-CANCELADA", 40, "Cancelado", "01/07/2026"])
            workbook.save(simphony)

            workbook = Workbook()
            workbook.active.append(
                ["Chave", "ValorContabil", "DataDocumento", "Cancelado"]
            )
            workbook.save(fiscal)

            workbook = Workbook()
            workbook.active.append(
                ["Chave de acesso", "Valor R$", "Data de emissão"]
            )
            workbook.save(sefaz)

            result = reconcile([simphony, fiscal, sefaz])

            self.assertEqual(result.rows[0].status, "Conciliado: cancelado")
            self.assertEqual(result.rows[0].simphony_status, "Cancelado")
            self.assertEqual(result.rows[0].difference, Decimal())
            self.assertTrue(result.rows[0].comparable)
