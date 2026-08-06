from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook

from automations.notas_entrada_atrasadas import analyze, note_status


class LateEntryNotesTest(TestCase):
    def test_marks_late_at_allowed_limit(self):
        self.assertEqual(note_status(10, "CE"), (11, "Alerta"))
        self.assertEqual(note_status(11, "CE"), (11, "Em atraso"))
        self.assertEqual(note_status(29, "SP"), (30, "Alerta"))
        self.assertEqual(note_status(30, "SP"), (30, "Alerta"))
        self.assertEqual(note_status(31, "SP"), (30, "Em atraso"))

    def test_rounds_elapsed_time_to_nearest_day(self):
        from automations.notas_entrada_atrasadas import rounded_days

        emission = datetime(2026, 1, 1, 12)
        self.assertEqual(rounded_days(emission, datetime(2026, 1, 12, 0)), 11)
        self.assertEqual(rounded_days(emission, datetime(2026, 1, 11, 23)), 10)

    def test_matches_received_note_by_company_and_key(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            manifesto = root / "manifesto.xlsx"
            detalhe = root / "detalhe.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Empresa",
                    "Chave Manifesto",
                    "Data Emissão",
                    "Estado",
                    "Fornecedor",
                ]
            )
            sheet.append(
                [
                    "CARMEL TAÍBA",
                    "CHAVE-1",
                    date(2026, 1, 1),
                    "bandeira-ceara.png",
                    "FORNECEDOR",
                ]
            )
            workbook.save(manifesto)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Empresa", "Chave", "Data de Entrada"])
            sheet.append(["CHARME HOSPEDAGEM", "CHAVE-1", date(2026, 1, 2)])
            workbook.save(detalhe)

            row = analyze(
                [manifesto, detalhe], reference_date=date(2026, 1, 20)
            ).rows[0]

            self.assertEqual(row.company, "CARMEL TAÍBA")
            self.assertIsNone(row.entry_date)
            self.assertEqual(row.launch_status, "Não lançada")
            self.assertEqual(row.status, "Em atraso")

    def test_keeps_manifest_row_with_empty_optional_trailing_columns(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            manifesto = root / "manifesto.xlsx"
            detalhe = root / "detalhe.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Empresa",
                    "Chave Manifesto",
                    "Data EmissÃ£o",
                    "Estado",
                    "Fornecedor",
                ]
            )
            sheet.append(["CARMEL TAÃBA", "CHAVE-SEM-CADASTRO", date(2026, 1, 1)])
            workbook.save(manifesto)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Empresa", "Chave", "Data de Entrada"])
            workbook.save(detalhe)

            rows = analyze(
                [manifesto, detalhe], reference_date=date(2026, 2, 1)
            ).rows

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].key, "CHAVE-SEM-CADASTRO")
            self.assertEqual(rows[0].state, "N/I")
            self.assertEqual(rows[0].supplier, "")
            self.assertEqual(rows[0].status, "Em atraso")
